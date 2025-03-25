from django.shortcuts import render, redirect,get_object_or_404,HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse
import os
import pandas as pd
from datetime import datetime
from django.core.exceptions import PermissionDenied
from django.core.files.storage import default_storage
from .models import LibraryResource, UserProfile,Department,ICTFacility,EContentDevelopment,Teacher,Expenditure,TeacherAward,ResearchGrant,Investigator,AwardRecognition,Patent,PhDAward,DemandRatio,ResearchPaper,BookChapter,AdmittedStudent,TeacherServingPost,FullTimeTeacher,TeacherAgainstSanctionedPost,Programme,ValueAddedCourse
from .models import Course,StudentProject
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate

def user_login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, "Invalid email address!")
            return redirect("user_login")

        # Authenticate user
        auth_user = authenticate(request, username=user.username, password=password)

        if auth_user is not None and auth_user.is_active:
            login(request, auth_user)
            messages.success(request, "Login successful!")

            # Check if UserProfile exists
            user_profile = UserProfile.objects.filter(user=user).first()
            if not user_profile:
                messages.error(request, "User profile not found. Contact Admin.")
                logout(request)
                return redirect("user_login")

            # Redirect based on role
            if user_profile.is_department_staff:
                return redirect("department_dashboard")
            elif user_profile.is_scholar:
                return redirect("scholar_dashboard")
            elif user_profile.is_teacher:
                return redirect("staff_dashboard")  # Teachers -> Staff Dashboard

            # If no role is assigned, logout user
            messages.error(request, "You are not authorized to access this system.")
            logout(request)
            return redirect("user_login")

        else:
            messages.error(request, "Invalid credentials!")
            return redirect("user_login")

    return render(request, "login.html")


@login_required
def department_dashboard(request):
    try:
        profile = request.user.userprofile
        if not profile.is_department_staff:
            messages.error(request, "You are not authorized to access this page.")
            return redirect("login")
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect("login")

    return render(request, "department_dashboard.html")

@login_required
def staff_dashboard(request):
    try:
        profile = request.user.userprofile
        if not profile.is_teacher:
            messages.error(request, "You are not authorized to access this page.")
            return redirect("login")
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect("login")

    return render(request, "staff_dashboard.html")

@login_required
def scholar_dashboard(request):
    try:
        profile = request.user.userprofile
        if not profile.is_scholar:
            messages.error(request, "You are not authorized to access this page.")
            return redirect("login")
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect("login")

    return render(request, "scholar_dashboard.html")



@login_required
def department_dashboard(request):
    # Ensure UserProfile exists and check `is_department_staff`
    if not hasattr(request.user, "userprofile") or not request.user.userprofile.is_department_staff:
        messages.error(request, "You are not authorized to access this page.")
        return redirect("home")
    
    return render(request, "department_dashboard.html")


@login_required
def scholar_dashboard(request):
    # Ensure UserProfile exists and check `is_scholar`
    if not hasattr(request.user, "userprofile") or not request.user.userprofile.is_scholar:
        messages.error(request, "You are not authorized to access this page.")
        return redirect("home")
    
    return render(request, "scholar_dashboard.html")

def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect("login")  # Redirect to login page after logout

@login_required
def home_view(request):
    return render(request, "home.html",  {'disable_filter': True})

def criterion_page(request, criterion_number):
    request.session['criterion_id'] = criterion_number
    return render(request, 'criterion.html', {'criterion_id': criterion_number})

def user_is_library_or_superuser(user):
    """Check if user is a superuser or belongs to the Library department."""
    if user.is_superuser:
        return True  # Admin can access everything
    try:
        return user.userprofile.is_library_user()
    except UserProfile.DoesNotExist:
        return False

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.core.paginator import Paginator
import openpyxl
from .models import LibraryResource, Department

@login_required
def library_resources(request):
    """
    List, filter, and export Library Resources:
    - Superusers see all resources.
    - Library department staff see only their department’s resources.
    - Teachers and unauthorized users are restricted.
    """
    current_year = datetime.now().year
    years = list(range(2000, current_year + 1))
    # Fetch the user and their profile
    user = request.user
    user_profile = getattr(user, 'userprofile', None)

    # Access Control
    if user.is_superuser:
        # Superusers see all data
        resources = LibraryResource.objects.all()
        departments = Department.objects.all()
    elif user_profile and user_profile.department and user_profile.department.department_name == "Library":
        # Library staff see their department's resources
        resources = LibraryResource.objects.filter(department=user_profile.department)
        departments = Department.objects.filter(department_code=user_profile.department.department_code)
    else:
        # Unauthorized users are redirected
        messages.error(request, "You are not authorized to view this page.")
        return redirect("home")

    # Apply Filters
    selected_department = request.GET.get('department')
    year_from = request.GET.get('year_from')
    year_to = request.GET.get('year_to')
    resource_name = request.GET.get('resource_name')

    if selected_department:
        resources = resources.filter(department__id=selected_department)
    if year_from and year_to:
        resources = resources.filter(academic_year__range=(year_from, year_to))
    elif year_from:
        resources = resources.filter(academic_year__gte=year_from)
    elif year_to:
        resources = resources.filter(academic_year__lte=year_to)
    if resource_name:
        resources = resources.filter(resource_name=resource_name)

    # Export to Excel
    if 'export_excel' in request.GET:
        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Library Resources"

        # Add headers
        sheet.append(['Academic Year', 'Resource Name', 'Expenditure on e-Journals/Books',
                      'Expenditure on Other e-Resources', 'Total Expenditure', 'Department'])

        # Add data rows
        for resource in resources:
            sheet.append([
                resource.academic_year,
                resource.resource_name,
                resource.expenditure_journals,
                resource.expenditure_other_resources,
                resource.total_expenditure,
                resource.department.department_name,
            ])

        # Return response as an Excel file
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=LibraryResources.xlsx'
        workbook.save(response)
        return response

    # Pagination
    paginator = Paginator(resources, 10)  # Show 10 resources per page
    page_number = request.GET.get('page')
    resources = paginator.get_page(page_number)

    # Context for rendering
    context = {
        "resources": resources,
        "departments": departments,
        "selected_department": selected_department,
        "year_from": year_from,
        "year_to": year_to,
        "resource_name": resource_name,
        "years": years,
    }

    return render(request, 'Forms/library_resources.html', context)

@login_required
def download_library_resources_excel(request):
    """Export Library Resources data to an Excel file with authorization checks."""

    user = request.user

    # **Check if user has a UserProfile**
    try:
        user_profile = user.userprofile
    except UserProfile.DoesNotExist:
        return HttpResponse("Unauthorized: No user profile found", status=403)

    # **Authorization Logic**
    if user.is_superuser:
        resources = LibraryResource.objects.all()
    elif user_profile.is_department_staff:
        if user_profile.department:
            resources = LibraryResource.objects.filter(department=user_profile.department)
        else:
            return HttpResponse("Unauthorized: No department assigned", status=403)
    else:
        return HttpResponse("Unauthorized: Insufficient permissions", status=403)

    # **Filtering Logic**
    year_from = request.GET.get("year_from")
    year_to = request.GET.get("year_to")
    resource_name = request.GET.get("resource_name")

    if year_from and year_to:
        resources = resources.filter(academic_year__range=(year_from, year_to))
    elif year_from:
        resources = resources.filter(academic_year__gte=year_from)
    elif year_to:
        resources = resources.filter(academic_year__lte=year_to)
    if resource_name:
        resources = resources.filter(resource_name=resource_name)

    # **Create an Excel Workbook**
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Library Resources"

    # **Header Row**
    headers = ["Academic Year", "Resource Name", "Expenditure on e-Journals/Books",
               "Expenditure on Other e-Resources", "Total Expenditure", "Department", "Downloaded By"]
    sheet.append(headers)

    # **Add Data Rows**
    for resource in resources:
        sheet.append([
            resource.academic_year,
            resource.resource_name,
            resource.expenditure_journals,
            resource.expenditure_other_resources,
            resource.total_expenditure,
            resource.department.department_name if resource.department else "N/A",
            f"{user.username} ({getattr(user_profile, 'position', 'N/A')})",  # Capture username & position
        ])

    # **Prepare HTTP Response**
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="LibraryResources.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def download_library_resource_pdf(request, resource_id):
    """
    Generate and download a PDF for Library Resource details.
    """
    resource = get_object_or_404(LibraryResource, id=resource_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="library_resource_{resource.id}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Background Color
    pdf.setFillColorRGB(0.95, 0.95, 0.95)  # Light Gray
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # Logo
    logo_path = finders.find("images/logo2.png")  # Replace with your logo path
    if not logo_path:
        raise OSError("Logo file not found in static directory!")
    logo = ImageReader(logo_path)

    # Header with Blue Background
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)

    # Header Text
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Library Resource Details")
    pdf.drawImage(logo, 50, height - 85, width=140, height=50, mask="auto")

    # Table Data
    data = [
        ["Academic Year", resource.academic_year],
        ["Resource Name", resource.resource_name],
        ["Expenditure on e-Journals/Books", f"₹{resource.expenditure_journals}"],
        ["Expenditure on Other e-Resources", f"₹{resource.expenditure_other_resources}"],
        ["Total Library Expenditure", f"₹{resource.total_expenditure}"],
        ["Department", resource.department.department_name],
    ]

    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # Draw Table in Center
    table_x = (width - 480) / 2
    table_y = height - 280
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # Footer with Blue Background
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)

    # Footer Text
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # Curved Double Line Border
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response

@login_required
def library_resource_detail(request, resource_id):
    """
    Show details for a specific Library Resource.
    """
    resource = get_object_or_404(LibraryResource, id=resource_id)
    return render(request, 'viewData/library_resource_detail.html', {'resource': resource})

@login_required
def update_library_resource(request, resource_id):
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("library_resources")  # Redirect to your library resources page

    resource = get_object_or_404(LibraryResource, id=resource_id)

    try:
        # Update fields with form data
        resource.academic_year = request.POST.get("academic_year", resource.academic_year)
        resource.resource_name = request.POST.get("resource_name", resource.resource_name)
        resource.expenditure_journals = float(request.POST.get("expenditure_journals", resource.expenditure_journals))
        resource.expenditure_other_resources = float(request.POST.get("expenditure_other_resources", resource.expenditure_other_resources))
        resource.total_expenditure = float(request.POST.get("total_expenditure", resource.total_expenditure))

        # Handle document upload if provided
        if "document" in request.FILES:
            resource.document = request.FILES["document"]

        resource.save()
        messages.success(request, "Library resource updated successfully.")
        return redirect("library_resources") 

    except ValueError as e:
        messages.error(request, f"Invalid data provided: {str(e)}")
        return redirect("library_resources")

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect("library_resources")


def validate_pdf(file):
    """Checks if the uploaded file is a PDF and does not exceed 5MB."""
    if file:
        if not file.name.lower().endswith(".pdf"):
            return "Only PDF files are allowed."
        if file.size > 5 * 1024 * 1024:  # 5MB limit
            return "File size should not exceed 5MB."
    return None

@login_required
def add_library_resource(request):
    """Handles adding a new library resource to the database."""
    if request.method == "POST":
        academic_year = request.POST.get("academicYear")
        resource_name = request.POST.get("resourceName")
        expenditure_journals = request.POST.get("expenditureJournals")
        expenditure_other_resources = request.POST.get("expenditureOthers")
        total_expenditure = request.POST.get("totalExpenditure")
        document = request.FILES.get("document")
        criterion_id = request.session.get('criterion_id')
        # Validate PDF file
        error = validate_pdf(document)
        if error:
            messages.error(request, error)
            return redirect("add_library_resource")

        # Get the Library department instance
        try:
            library_department = Department.objects.get(department_name="Library")
        except Department.DoesNotExist:
            messages.error(request, "Library department not found.")
            return redirect("add_library_resource")

        # Convert expenditure values to float
        try:
            expenditure_journals = float(expenditure_journals)
            expenditure_other_resources = float(expenditure_other_resources)
            total_expenditure = float(total_expenditure)
        except ValueError:
            messages.error(request, "Invalid expenditure values entered.")
            return redirect("add_library_resource")

        # Save resource data in the database
        LibraryResource.objects.create(
            academic_year=academic_year,
            resource_name=resource_name,
            expenditure_journals=expenditure_journals,
            expenditure_other_resources=expenditure_other_resources,
            total_expenditure=total_expenditure,
            document=document,
            department=library_department
        )

        messages.success(request, "Library resource added successfully!")
        return redirect("library_resources")

    return render(request, "AddData/add_library_resource.html", {'disable_filter': True})

def delete_library_resource(request, resource_id):
    resource = get_object_or_404(LibraryResource, id=resource_id)
    resource.delete()
    return redirect('library_resources')  # Redirect to the main page after deletion
@login_required
def ict_facility_list(request):
    """
    List and filter ICT Facilities:
    - Superusers see all data.
    - Department staff see data related to their department.
    - Teachers see only their related department data.
    """

    # Fetch the user and their profile
    user = request.user
    user_profile = getattr(user, 'userprofile', None)

    # Fetch filter values from GET request
    department_id = request.GET.get('department', '')
    room_type = request.GET.get('room_type', '')
    ict_facility = request.GET.get('ict_facility', '')

    # Initialize the facilities queryset and restrict access based on roles
    if user.is_superuser:
        # Superusers see all facilities
        facilities = ICTFacility.objects.all()
        departments = Department.objects.all()
    elif user_profile and user_profile.is_department_staff:
        # Department staff see only their department's facilities
        facilities = ICTFacility.objects.filter(department=user_profile.department)
        departments = Department.objects.filter(department_code=user_profile.department.department_code)
    elif user_profile and user_profile.is_teacher:
        # Teachers see only their department's facilities
        if user_profile.department:
            facilities = ICTFacility.objects.filter(department=user_profile.department)
            departments = Department.objects.filter(department_code=user_profile.department.department_code)
        else:
            facilities = ICTFacility.objects.none()
            departments = Department.objects.none()
    else:
        # For users without a proper profile, return no data
        facilities = ICTFacility.objects.none()
        departments = Department.objects.none()

    # Apply additional filters if provided in the GET request
    if department_id:
        facilities = facilities.filter(department_id=department_id)
    if room_type:
        facilities = facilities.filter(room_type__icontains=room_type)  # Case-insensitive matching
    if ict_facility:
        facilities = facilities.filter(ict_facility__icontains=ict_facility)

    # Add pagination for scalability
    from django.core.paginator import Paginator
    paginator = Paginator(facilities, 10)  # Show 10 facilities per page
    page_number = request.GET.get('page')
    facilities = paginator.get_page(page_number)

    # Context for rendering the template
    context = {
        'facilities': facilities,
        'departments': departments,
        'selected_department': department_id,
        'selected_room_type': room_type,
        'selected_ict_facility': ict_facility,
    }

    return render(request, 'Forms/ict_facility_list.html', context)

@login_required
def download_ict_facilities(request):
    """Generate and download an Excel file containing ICT facilities details based on user role."""

    user = request.user

    # **Check if user has a UserProfile**
    try:
        user_profile = user.userprofile
    except UserProfile.DoesNotExist:
        return HttpResponse("Unauthorized: No user profile found", status=403)

    # **Authorization Logic**
    if user.is_superuser:
        facilities = ICTFacility.objects.all()
    elif user_profile.is_department_staff:
        if user_profile.department:
            facilities = ICTFacility.objects.filter(department=user_profile.department)
        else:
            return HttpResponse("Unauthorized: No department assigned", status=403)
    else:
        return HttpResponse("Unauthorized: Insufficient permissions", status=403)

    # **Prepare Data**
    data = []
    for facility in facilities:
        data.append({
            "Department Code": facility.department.department_code,
            "Department Name": facility.department.department_name,
            "Room Type": facility.room_type,
            "Room No": facility.room_no,
            "Type of ICT Facility": facility.ict_facility,
            "Geo-Tagged Photo": facility.geo_tagged_photo.url if facility.geo_tagged_photo else "No Image",
            "Master Time Table": facility.master_timetable.url if facility.master_timetable else "No Timetable",
            "Downloaded By": f"{user.username} ({getattr(user_profile, 'position', 'N/A')})",  # Capture username & position
        })

    # **Convert to DataFrame**
    df = pd.DataFrame(data)

    # **Create HTTP response with Excel file**
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ICT_Facilities.xlsx"'

    # **Write to Excel file**
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='ICT Facilities')

    return response

from reportlab.pdfgen import canvas

def facility_detail(request, facility_id):
    facility = get_object_or_404(ICTFacility, id=facility_id)
    return render(request, "viewData/facility_detail.html", {"facility": facility})

def download_facility_pdf(request, facility_id):
    facility = get_object_or_404(ICTFacility, id=facility_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="facility_{facility.id}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4  

    # **Background Color for Entire Page**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)  # Light Gray Background
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # ✅ Find Logo File
    logo_path = finders.find("images/logo2.png")
    if not logo_path:
        raise OSError("Logo file not found in static directory!")

    logo = ImageReader(logo_path)

    # **Header Section with Blue Background**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)  # ✅ Rounded Header

    # **Header Text**
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "ICT Facility Details")

    pdf.drawImage(logo, 50, height - 85, width=140, height=50, mask="auto")

    # **Table Data**
    data = [
        ["Department Code", facility.department.department_code],
        ["Department Name", facility.department.department_name],
        ["Room Type", facility.room_type],
        ["Room No", facility.room_no],
        ["Type of ICT Facility", facility.ict_facility]
    ]

    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),  # ✅ Blue Borders
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),   # ✅ Outer Blue Border
    ]))

    # **Draw Table in Center**
    table_x = (width - 480) / 2  
    table_y = height - 280  
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section with Blue Background**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)  # ✅ Rounded Footer

    # **Footer Text Without Logo**
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Curved Double Line Border for A4 Page**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)  # ✅ Outer Border
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)  # ✅ Inner Border

    pdf.showPage()
    pdf.save()
    
    return response

@login_required
def add_ict_facility(request):
    """ Add a new ICT Facility """
    criterion_id = request.session.get('criterion_id')
    if request.method == "POST":
        department_code = request.POST.get("department", "").strip()
        room_type = request.POST.get("room_type", "").strip()
        room_no = request.POST.get("room_no", "").strip()
        ict_facility = request.POST.get("ict_facility", "").strip()
        geo_tagged_photo = request.FILES.get("geo_tagged_photo")  # Accepts file
        master_timetable = request.FILES.get("master_timetable")  # Accepts file
        if not (room_type and room_no and ict_facility and geo_tagged_photo and master_timetable):
            messages.error(request, "All fields are required!")
            return redirect("add_facility")

        department = get_object_or_404(Department, department_code=department_code)

        ICTFacility.objects.create(
            department=department,
            room_type=room_type,
            room_no=room_no,
            ict_facility=ict_facility,
            geo_tagged_photo=geo_tagged_photo,
            master_timetable=master_timetable,
        )

        messages.success(request, "New ICT Facility added successfully!")
        return redirect("ict_facility_list")

    departments = Department.objects.all()
    return render(request, "AddData/add_ict_facility.html", {"departments": departments,'criterion_id':criterion_id, 'disable_filter': True})

@login_required
def update_facility(request, facility_id):
    """ Update ICT Facility using POST data """
    facility = get_object_or_404(ICTFacility, id=facility_id)

    if request.method == "POST":
        room_type = request.POST.get("room_type", "").strip()
        room_no = request.POST.get("room_no", "").strip()
        ict_facility = request.POST.get("ict_facility", "").strip()

        # Handling file uploads correctly
        geo_tagged_photo = request.FILES.get("geo_tagged_photo")
        master_timetable = request.FILES.get("master_timetable")

        # Validate required fields (excluding optional files)
        if not (room_type and room_no and ict_facility):
            messages.error(request, "Room Type, Room No, and ICT Facility are required!")
            return redirect("update_facility", facility_id=facility.id)

        # Update facility details
        facility.room_type = room_type
        facility.room_no = room_no
        facility.ict_facility = ict_facility

        # Only update file fields if a new file is uploaded
        if geo_tagged_photo:
            facility.geo_tagged_photo = geo_tagged_photo
        if master_timetable:
            facility.master_timetable = master_timetable

        facility.save()
        messages.success(request, "ICT Facility updated successfully!")
        return redirect("ict_facility_list")

    return render(request, "Forms/ict_facility_list.html", {"facility": facility})

@login_required
def view_econtent_detail(request, record_id):
    econtent = get_object_or_404(EContentDevelopment, id=record_id)
    return render(request, "viewData/econtent_detail.html", {"econtent": econtent})
@login_required
def download_econtent_excel(request):
    """Generate and download an Excel file for the logged-in user's E-Content records."""

    user = request.user

    # **Check if the user has a UserProfile**
    try:
        user_profile = user.userprofile
    except UserProfile.DoesNotExist:
        return HttpResponse("Unauthorized: No user profile found", status=403)

    # **Determine access level**
    if user.is_superuser:
        records = EContentDevelopment.objects.all()
    elif user_profile.is_teacher:
        records = EContentDevelopment.objects.filter(teacher__user_profile=user_profile)
    elif user_profile.is_department_staff:
        if user_profile.department:
            records = EContentDevelopment.objects.filter(teacher__user_profile__department=user_profile.department)
        else:
            return HttpResponse("Unauthorized: No department assigned", status=403)
    else:
        return HttpResponse("Unauthorized: Insufficient permissions", status=403)

    # **Create Excel file**
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E-Content Records"

    # **Headers**
    headers = ["Module Name", "Platform", "Launch Date", "Teacher (Position)", "Department", "Facilities", "Document Available", "Video Link"]
    ws.append(headers)

    # **Populate rows**
    for record in records:
        teacher = record.teacher
        teacher_name = teacher.user_profile.user.username if teacher else "N/A"
        teacher_position = f" ({teacher.position})" if teacher and teacher.position else ""
        teacher_full_name = f"{teacher_name}{teacher_position}"

        ws.append([
            record.module_name,
            record.platform,
            record.launch_date.strftime("%Y-%m-%d") if record.launch_date else "N/A",
            teacher_full_name,
            teacher.user_profile.department.department_name if teacher and teacher.user_profile.department else "N/A",
            record.facility_available,
            "Yes" if record.document_link else "No",
            record.video_link if record.video_link else "N/A",
        ])

    # **Prepare HTTP response**
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="E_Content_Records.xlsx"'

    wb.save(response)
    return response

@login_required
def delete_ict_facility(request, facility_id):
    facility = get_object_or_404(ICTFacility, id=facility_id)
    facility.delete()
    return redirect('ict_facility_list')

def add_econtent(request):
    criterion_id = request.session.get('criterion_id')
    if request.method == 'POST':
        name_teacher = request.POST.get('nameTeacher')  # Get Teacher ID or name
        module_name = request.POST.get('moduleName')
        platform = request.POST.get('platform')
        launch_date = request.POST.get('launchDate')
        document_link = request.FILES.get('docFile')
        facility_list = request.POST.get('facilityList')  # Comma-separated list
        video_link = request.POST.get('videoLink')
     

        # Convert facility list to JSON format
    
        # Retrieve the Teacher object
        teacher_obj = None
        if name_teacher.isdigit():  # If Teacher ID is selected from dropdown
            teacher_obj = Teacher.objects.get(id=int(name_teacher))
        else:  # If teacher's name is auto-filled
            teacher_obj = Teacher.objects.filter(user__username=name_teacher).first()

        # Check if teacher exists
        if not teacher_obj:
            messages.error(request, "Invalid teacher selection.")
            return redirect('add_econtent')

        # Save the new EContentDevelopment entry
        econtent = EContentDevelopment(
            teacher=teacher_obj,
            module_name=module_name,
            platform=platform,
            launch_date=launch_date,
            document_link=document_link,
            facility_available=facility_list,
            video_link=video_link
        )
        econtent.save()

        messages.success(request, "E-Content record added successfully.")
        return redirect('show_econtent')  # Redirect to list view

    # Fetch required data for rendering form
    teachers = Teacher.objects.all()
    user_role = request.user.groups.first().name if request.user.groups.exists() else "guest"
    
    logged_in_teacher = None
    if user_role == "teacher":
        try:
            logged_in_teacher = Teacher.objects.get(user=request.user)
        except Teacher.DoesNotExist:
            messages.error(request, "You are not associated with a teacher record.")

    return render(request, 'AddData/add_econtent.html', {
        'teachers': teachers,
        'user_role': user_role,
        'criterion_id':criterion_id,
        'logged_in_teacher': logged_in_teacher.user.username if logged_in_teacher else "",
        'disable_filter': True
    })

@login_required
def show_econtent(request):
    """Display E-Content records based on user roles with filters."""
    
    user = request.user
    current_year = datetime.now().year
    year_range = range(2000, current_year + 1)  # Generate years from 2000 to current year
    is_superuser = user.is_superuser
    user_profile = get_object_or_404(UserProfile, user=user) if not is_superuser else None

    # Initialize filters
    selected_teacher = request.GET.get('teacher')
    selected_department = request.GET.get('department')
    from_year = request.GET.get('from_year')
    to_year = request.GET.get('to_year')

    # Superuser: See all records, departments, and teachers
    if is_superuser:
        records = EContentDevelopment.objects.all()
        departments = Department.objects.all()
        teachers = Teacher.objects.all()

    # Department Staff: View all teachers in their department
    elif user_profile.is_department_staff and user_profile.department:
        department = user_profile.department
        records = EContentDevelopment.objects.filter(teacher__user_profile__department=department)
        departments = Department.objects.filter(department_code=department.department_code)
        teachers = Teacher.objects.filter(user_profile__department=department)

    # Teachers: See only their own records
    elif user_profile.is_teacher:
        teacher = get_object_or_404(Teacher, user_profile=user_profile)
        records = EContentDevelopment.objects.filter(teacher=teacher)
        departments = Department.objects.filter(department_code=user_profile.department.department_code)
        teachers = None  # No teacher dropdown needed

    # Unauthorized users: No access
    else:
        records = EContentDevelopment.objects.none()
        departments = Department.objects.none()
        teachers = None

    # Apply Filters
    if from_year:
        records = records.filter(launch_date__year__gte=from_year)
    if to_year:
        records = records.filter(launch_date__year__lte=to_year)
    if selected_teacher and teachers:
        records = records.filter(teacher_id=selected_teacher)
    if selected_department and departments:
        records = records.filter(teacher__user_profile__department__department_code=selected_department)

    context = {
        "records": records,  
        "departments": departments,  
        "teachers": teachers,  
        "year_range": year_range,
        "selected_teacher": selected_teacher,
        "selected_department": selected_department,
        "from_year": from_year,
        "to_year": to_year,
        "is_superuser": is_superuser,
        "is_department_staff": getattr(user_profile, 'is_department_staff', False),
    }

    return render(request, 'Forms/econtent_list.html', context)
def download_econtent_pdf(request, econtent_id):
    """Generate and download a PDF for a single viewed E-Content Development record."""
    
    econtent = get_object_or_404(EContentDevelopment, id=econtent_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="econtent_{econtent.id}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4  

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # **Find Logo File**
    try:
        logo_path = finders.find("images/logo2.png")
        if logo_path:
            logo = ImageReader(logo_path)
            pdf.drawImage(logo, 50, height - 85, width=140, height=50, mask="auto")
    except Exception as e:
        print(f"Error loading logo: {e}")

    # **Header Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "E-Content Development Details")

    # **Retrieve Safe Data**
    teacher_name = econtent.teacher.user_profile.user.username  if econtent.teacher else "N/A"
    position= econtent.teacher.position if econtent.teacher else "N/A"
    department_name = econtent.teacher.user_profile.department.department_name if econtent.teacher and econtent.teacher.user_profile.department else "N/A"
    document_status = "Available" if econtent.document_link else "Not Available"
    video_status = econtent.video_link if econtent.video_link else "N/A"

    # **Table Data**
    data = [
        ["Module Name", econtent.module_name],
        ["Platform", econtent.platform],
        ["Launch Date", econtent.launch_date.strftime("%d-%m-%Y") if econtent.launch_date else "N/A"],
        ["Teacher", teacher_name + " (" + position + ")"],
        ["Department", department_name],
        ["Facilities", econtent.facility_available],
        ["Document", document_status],
        ["Video Link", video_status]
    ]

    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    table_x = (width - 480) / 2  
    table_y = height - 280  
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)

    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Curved Double Line Border**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()
    
    return response



@login_required
def edit_econtent(request, record_id):
    """Edit E-Content: Only the associated teacher or an admin can edit."""
    record = get_object_or_404(EContentDevelopment, id=record_id)

    if not request.user.is_superuser and request.user != record.teacher.user_profile.user:
        messages.error(request, "You are not authorized to edit this record.")
        return redirect('show_econtent')

    if request.method == "POST":
        record.module_name = request.POST.get("module_name")
        record.platform = request.POST.get("platform")

        launch_date_str = request.POST.get("launch_date")
        if launch_date_str:
            try:
                record.launch_date = datetime.strptime(launch_date_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
                return redirect('show_econtent')

        record.facility_available = request.POST.get("facility_available") 
        record.video_link = request.POST.get("video_link")

        # Handle file upload (retain existing file if no new file is uploaded)
        if 'document_link' in request.FILES:
            record.document_link = request.FILES['document_link']

        record.save()
        messages.success(request, "E-Content updated successfully!")
        return redirect('show_econtent')

    return redirect('show_econtent')

@login_required
def delete_econtent(request, record_id):
    """Delete E-Content: Only the associated teacher or an admin can delete."""
    record = get_object_or_404(EContentDevelopment, id=record_id)
    record.delete()
    messages.success(request,"E-Content deleted successfully! ")
    return redirect('show_econtent')

@login_required
def expenditure_list(request):
    """Show Expenditure records with filters based on user roles (Admins see all, department users see their own)."""

    # ✅ Get filter parameters from the request
    selected_department = request.GET.get('department', None)  # For department filtering
    year_from = request.GET.get('year_from', None)  # Year From Filter
    year_to = request.GET.get('year_to', None)      # Year To Filter

    # Determine the records based on the user's role
    if request.user.is_superuser:
        # Superuser can access all expenditure records
        records = Expenditure.objects.all()

        # Apply department filter if provided
        if selected_department:
            records = records.filter(department__department_code=selected_department)
    else:
        # Non-superuser (department users) can only access their department's data
        user_profile = get_object_or_404(UserProfile, user=request.user)
        if user_profile.department:
            records = Expenditure.objects.filter(department=user_profile.department)
        else:
            # No department assigned, no access to expenditure data
            records = Expenditure.objects.none()

    # Apply year filters (if provided)
    if year_from and year_to:
        records = records.filter(year__gte=year_from, year__lte=year_to)
    elif year_from:
        records = records.filter(year__gte=year_from)
    elif year_to:
        records = records.filter(year__lte=year_to)

    # Fetch departments for filtering dropdown
    if request.user.is_superuser:
        # Superuser sees all departments
        departments = Department.objects.all()
    else:
        # Department users see their own department only
        departments = Department.objects.filter(department_code=user_profile.department.department_code)

    # ✅ Context for the template
    context = {
        "departments": departments,
        "selected_department": selected_department,
        "year_from": year_from,
        "year_to": year_to,
        "expenditures": records,  # Pass the filtered expenditure records
    }

    return render(request, 'Forms/expenditure_list.html', context)
@login_required
def download_expenditure_excel(request):
    """Export expenditure data to an Excel file."""
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Expenditures"

    # Add headers
    headers = [
        "Year", "Department", "Budget Allocated",
        "Expenditure on Infrastructure", "Expenditure on Maintenance of Academic Facilities",
        "Expenditure on Maintenance of Physical Facilities", "Total Expenditure Excluding Salary"
    ]
    sheet.append(headers)

    # Add data rows
    expenditures = Expenditure.objects.all()
    for expenditure in expenditures:
        sheet.append([
            expenditure.year,
            expenditure.department.department_name,
            expenditure.budget_allocated,
            expenditure.expenditure_infra,
            expenditure.academic_facilities,
            expenditure.physical_facilities,
            expenditure.total_expenditure
        ])

    # Set up the response for downloading
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Expenditures.xlsx"'
    workbook.save(response)

    return response

@login_required
def view_expenditure_detail(request, expenditure_id):
    """
    Display the details of a specific expenditure record.
    """
    expenditure = get_object_or_404(Expenditure, id=expenditure_id)  # Get the record or return 404
    context = {'expenditure': expenditure }
    return render(request, 'viewData/expenditure_detail.html', context)
@login_required
def download_expenditure_pdf(request, expenditure_id):
    """
    Generate and download a PDF for a specific Expenditure record.
    """
    # Get the expenditure record
    expenditure = get_object_or_404(Expenditure, id=expenditure_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="expenditure_{expenditure.year}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # **Add Logo**
    try:
        logo_path = finders.find("images/logo2.png")  # Ensure the path is correct in static files
        if logo_path:
            logo = ImageReader(logo_path)
            pdf.drawImage(logo, 50, height - 85, 140, 50, mask="auto")
        else:
            print("Logo file not found.")
    except Exception as e:
        print(f"Error loading logo: {e}")

    # **Header Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Expenditure Details")

    # **Table Data**
    department_name = expenditure.department.department_name if expenditure.department else "N/A"
    data = [
        ["Year", expenditure.year],
        ["Department", department_name],
        ["Budget Allocated", f"₹ {expenditure.budget_allocated:,}"],
        ["Expenditure on Infrastructure", f"₹ {expenditure.expenditure_infra:,}"],
        ["Expenditure on Academic Facilities", f"₹ {expenditure.academic_facilities:,}"],
        ["Expenditure on Physical Facilities", f"₹ {expenditure.physical_facilities:,}"],
        ["Total Expenditure Excluding Salary", f"₹ {expenditure.total_expenditure:,}"],
    ]

    # **Style Table**
    table = Table(data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # **Draw Table**
    table_x = (width - 500) / 2
    table_y = height - 300
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
    pdf.drawCentredString(width / 2, 30, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Curved Double Line Border**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response
 
@login_required
def edit_expenditure(request, record_id):
    """Edit Expenditure: Only the associated department user or an admin can edit."""
    expenditure = get_object_or_404(Expenditure, id=record_id)

    # Restrict normal users from editing other departments' expenditures
    if not request.user.is_superuser and expenditure.department != request.user.userprofile.department:
        messages.error(request, "You are not authorized to edit this record.")
        return redirect("expenditure_list")

    if request.method == "POST":
        try:
            # Fetch values safely and convert them to proper data types
            expenditure.year = int(request.POST.get("year", expenditure.year))
            expenditure.budget_allocated = float(request.POST.get("budget_allocated", 0))
            expenditure.expenditure_infra = float(request.POST.get("expenditure_infra", 0))
            expenditure.academic_facilities = float(request.POST.get("academic_facilities", 0))
            expenditure.physical_facilities = float(request.POST.get("physical_facilities", 0))
            expenditure.total_expenditure = float(request.POST.get("total_expenditure_excluding_salary", 0))

            # Ensure values are non-negative
            if any(value < 0 for value in [
                expenditure.budget_allocated,
                expenditure.expenditure_infra,
                expenditure.academic_facilities,
                expenditure.physical_facilities,
                expenditure.total_expenditure
            ]):
                messages.error(request, "Values cannot be negative.")
                return redirect("expenditure_list")

            # Save updated data
            expenditure.save()
            messages.success(request, "Expenditure record updated successfully.")
            return redirect("expenditure_list")

        except ValueError:
            messages.error(request, "Invalid input. Please enter valid numeric values.")
            return redirect("expenditure_list")

    return redirect("expenditure_list")

@login_required
def delete_expenditure(request, record_id):
    expenditure = get_object_or_404(Expenditure, id=record_id)

    # Restrict normal users from deleting other departments' expenditures
    if not request.user.is_superuser and expenditure.department != request.user.userprofile.department:
        messages.error(request, "Unauthorized access: You cannot delete this record.")
        return JsonResponse({"success": False, "error": "Unauthorized"}, status=403)

    try:
        expenditure.delete()
        messages.success(request, "Expenditure record deleted successfully.")
        return redirect('expenditure_list')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('expenditure_list')

def add_expenditure(request):
    criterion_id = request.session.get('criterion_id')
    if request.method == "POST":
        #  Allow admins to proceed without a UserProfile
        if request.user.is_superuser:
            department = None  # Admins may not have a department
        else:
            try:
                user_profile = UserProfile.objects.get(user=request.user)  # Get logged-in user's profile
                department = user_profile.department
            except UserProfile.DoesNotExist:
                messages.error(request, "Your profile does not exist. Please contact the admin.")
                return redirect("add_expenditure")

        # Ensure a department is assigned (except for admins)
        if not department and not request.user.is_superuser:
            messages.error(request, "You do not have a department assigned.")
            return redirect("add_expenditure")

        try:
            year = int(request.POST.get("year", 0))
            budget = float(request.POST.get("budget", 0))
            expenditure = float(request.POST.get("expenditure", 0))
            total_expenditure = float(request.POST.get("total_expenditure", 0))
            academic_expenditure = float(request.POST.get("academic_expenditure", 0))
            physical_expenditure = float(request.POST.get("physical_expenditure", 0))

            #  Define the max limit based on DecimalField(max_digits=12, decimal_places=2)
            MAX_VALUE = 9999999999.99  

            # Check if any value exceeds the allowed limit
            if any(value > MAX_VALUE for value in [budget, expenditure, total_expenditure, academic_expenditure, physical_expenditure]):
                messages.error(request, "One or more values exceed the allowed limit.")
                return redirect("add_expenditure")

            # Save data if all values are within the limit
            Expenditure.objects.create(
                department=department if department else Department.objects.first(),  # Assign a default department if admin
                year=year,
                budget_allocated=budget,
                expenditure_infra=expenditure,
                total_expenditure=total_expenditure,
                academic_facilities=academic_expenditure,
                physical_facilities=physical_expenditure
            )

            messages.success(request, "Expenditure added successfully!")
            return redirect("expenditure_list")  # Redirect to the expenditure list page

        except ValueError:
            messages.error(request, "Invalid input! Please enter valid numbers.")
            return redirect("add_expenditure")

    return render(request, "AddData/add_expenditure.html",{'criterion_id':criterion_id, 'disable_filter': True})

def form5_view(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'Forms/form5.html',{'criterion_id':criterion_id})
def adddata5(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'AddData/adddata5.html',{'criterion_id':criterion_id, 'disable_filter': True})
def form6_view(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'Forms/form6.html',{'criterion_id':criterion_id})
def adddata6(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'AddData/adddata6.html',{'criterion_id':criterion_id,'disable_filter': True})
def form7_view(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'Forms/form7.html',{'criterion_id':criterion_id})
def adddata7(request):
    criterion_id = request.session.get('criterion_id')
    return render(request,'AddData/adddata7.html',{'criterion_id':criterion_id,'disable_filter': True})
def is_admin(user):
    return user.is_staff  # Assuming admin users have is_staff=True
@login_required
def teacher_awards_list(request):
    """Show Teacher Awards with filters (Admins see all, normal users see their own)."""

    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')

    # ✅ Determine records based on user role
    if is_admin(request.user):
        awards = TeacherAward.objects.select_related('teacher__user_profile').all()
        departments = Department.objects.all()

        # ✅ Apply department filter (for Admins)
        if selected_department:
            awards = awards.filter(teacher__user_profile__department__department_code=selected_department)

    else:
        # ✅ Normal users can only see their own awards
        user_profile = get_object_or_404(UserProfile, user=request.user)
        awards = TeacherAward.objects.select_related('teacher__user_profile').filter(
            teacher__user_profile=user_profile
        )

        # ✅ Normal users can only see their own department
        departments = Department.objects.filter(department_code=request.user.userprofile.department.department_code)

    return render(request, 'Forms/teacher_awards.html', {
        "awards": awards,
        "departments": departments, 
        "selected_department": selected_department,
        "criterion_id": criterion_id,
    })


@login_required
def update_teacher_award(request, award_id):
    """Allows only the owner or an admin to edit a teacher award entry."""
    award = get_object_or_404(TeacherAward, id=award_id)

    if not request.user.is_staff and award.teacher.user_profile != request.user.userprofile:
        messages.error(request, "Unauthorized access!")
        return redirect('teacher_awards_list')  

    if request.method == "POST":
        award.teacher_name = request.POST.get("teacher_name")
        award.award_name = request.POST.get("award_name")
        award.recognition_level = request.POST.get("recognition_level")
        award.year_of_award = request.POST.get("year_of_award")
        award.awarding_agency = request.POST.get("awarding_agency")
        award.save()

        messages.success(request, "Teacher award updated successfully!")
        return redirect('teacher_awards_list')  

    messages.error(request, "Invalid request!")
    return redirect('teacher_awards_list')  


@login_required
def delete_teacher_award(request, award_id):
    """Allows only the owner or an admin to delete a teacher award entry."""
    award = get_object_or_404(TeacherAward, id=award_id)

    if not request.user.is_staff and award.teacher.user_profile != request.user.userprofile:
        messages.error(request, "Unauthorized access!")
        return redirect('teacher_awards_list')

    award.delete()
    messages.success(request, "Teacher award deleted successfully!")
    return redirect('teacher_awards_list') 



@login_required
def add_teacher_award(request):
    user = request.user
    criterion_id = request.session.get('criterion_id')

    if request.method == "POST":
        if user.is_superuser:
            teacher_id = request.POST.get("teacher")  
            try:
                teacher = Teacher.objects.get(id=teacher_id)
            except Teacher.DoesNotExist:
                messages.error(request, "Selected teacher does not exist.")
                return redirect("add_teacher_award")
        else:
            try:
                user_profile = UserProfile.objects.get(user=user) 
                teacher = Teacher.objects.get(user_profile=user_profile) 
            except (UserProfile.DoesNotExist, Teacher.DoesNotExist):
                messages.error(request, "Your teacher profile does not exist.")
                return redirect("add_teacher_award")

        # Get form fields
        award_name = request.POST.get("award_name")
        recognition_level = request.POST.get("recognition_level")
        year_of_award = request.POST.get("award_year")
        awarding_agency = request.POST.get("award_agency")

        # Save award entry
        TeacherAward.objects.create(
            teacher=teacher,
            award_name=award_name,
            recognition_level=recognition_level,
            year_of_award=year_of_award,
            awarding_agency=awarding_agency
        )

        messages.success(request, "Teacher Award added successfully!")
        return redirect("teacher_awards_list")  # Redirect to awards list

    teachers = Teacher.objects.all() if user.is_superuser else None 
    return render(request, "AddData/add_teacher_award.html", {"teachers": teachers, "is_admin": user.is_superuser,'criterion_id':criterion_id,'disable_filter': True})
@login_required
def research_grants_list(request):
    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')  # Get department filter from request

    if request.user.is_superuser:
        departments = Department.objects.all()
        grants = ResearchGrant.objects.all()

        # Apply filtering if a department is selected
        if selected_department:
            grants = grants.filter(department__department_code=selected_department)
    else:
        try:
            department = request.user.profile.department  # Get user's department
            departments = [department]  # Convert to list for template
            grants = ResearchGrant.objects.filter(department=department)
        except AttributeError:
            department = None
            departments = []
            grants = ResearchGrant.objects.none()  # Handle missing profile

    return render(request, 'Forms/research_grants_list.html', {
        'grants': grants,
        'criterion_id': criterion_id,
        'departments': departments,
        'selected_department': selected_department  # Pass selected department to template
    })

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def add_research_grant(request):
    criterion_id = request.session.get('criterion_id')
    if request.method == "POST":
        scheme_name = request.POST.get("scheme_name")
        funding_agency = request.POST.get("funding_agency")
        grant_type = request.POST.get("grant_type")
        department_id = request.POST.get("department")  # Ensure this is the correct value
        year_of_award = request.POST.get("year_of_award")
        funds_provided = request.POST.get("funds_provided")
        duration = request.POST.get("duration")
        duration_unit = request.POST.get("duration_unit")
        investigators = request.POST.getlist("investor[]")

        # Determine the correct department
        if request.user.is_superuser:  # Admin can select department manually
            department = get_object_or_404(Department, department_code=department_id)
        else:  # Regular user should have their own department auto-filled
            department = request.user.userprofile.department

        # Create and save ResearchGrant instance
        grant = ResearchGrant.objects.create(
            scheme_name=scheme_name,
            funding_agency=funding_agency,
            grant_type=grant_type,
            department=department,
            year_of_award=year_of_award,
            funds_provided=funds_provided,
            duration=duration,
            duration_unit=duration_unit
        )

        # Save Investigators and link them to the grant
        investigator_objects = [
            Investigator.objects.get_or_create(name=name.strip())[0]
            for name in investigators if name.strip()
        ]
        grant.investigators.set(investigator_objects)

        return redirect("research_grants_list")

    # Fetch all departments for dropdown (Only admins need this)
    departments = Department.objects.all() if request.user.is_superuser else None

    # Get the logged-in user's department
    user_department = None
    if hasattr(request.user, 'userprofile') and request.user.userprofile.department:
        user_department = request.user.userprofile.department

    return render(request, "AddData/add_research_grant.html", {
        "departments": departments,
        "department": user_department,
        'criterion_id':criterion_id , # Auto-filled for users
        'disable_filter': True
    })

def update_grant(request, grant_id):
    if request.method == "POST":
        grant = get_object_or_404(ResearchGrant, id=grant_id)

        try:
            # Basic grant details
            grant.scheme_name = request.POST.get("scheme_name", "").strip()
            grant.funding_agency = request.POST.get("funding_agency", "").strip()
            grant.type = request.POST.get("type", "").strip()

            # Handle department correctly
            department_name = request.POST.get("department", "").strip()
            grant.department = get_object_or_404(Department, department_name=department_name)

            # Convert numerical fields safely
            grant.year_of_award = int(request.POST.get("year_of_award", 0))
            grant.funds_provided = float(request.POST.get("funds_provided", 0))
            grant.duration = int(request.POST.get("duration", 0))
            grant.duration_unit = request.POST.get("duration_unit", "").strip()

            # Handle multiple principal investigators
            investigator_names = request.POST.getlist("principal_investigator[]")  # Fetch all investigators
            grant.investigators.clear()  # Remove old investigators
            
            for name in investigator_names:
                name = name.strip()
                if name:  # Ensure it's not empty
                    investigator, _ = Investigator.objects.get_or_create(name=name)
                    grant.investigators.add(investigator)

            grant.save()
            messages.success(request, "Research grant updated successfully!")
        
        except ValueError as e:
            messages.error(request, f"Invalid input: {str(e)}")
        except Exception as e:
            messages.error(request, f"Error updating grant: {str(e)}")

    return redirect("research_grants_list")

def delete_grant(request, grant_id):
    grant = get_object_or_404(ResearchGrant, id=grant_id)

    try:
        grant.delete()
        messages.success(request, "Research grant deleted successfully!")
    except Exception as e:
        messages.error(request, f"Error deleting grant: {str(e)}")

    return redirect("research_grants_list")

@login_required
def award_list(request):
    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')
    """List all awards. Admins see all, users see only their department's awards."""
    if request.user.is_superuser:  
        awards = AwardRecognition.objects.all()  # Admin sees all records
    else:
        user_profile = get_object_or_404(UserProfile, user=request.user)
        awards = AwardRecognition.objects.filter(department=user_profile.department)  # Filter by user's department
    
    return render(request, 'Forms/award_list.html', {'awards': awards,'department':selected_department,'criterion_id':criterion_id})

@login_required
def update_award(request, award_id):
    """Update award details (only within the user's department unless admin)."""
    award = get_object_or_404(AwardRecognition, id=award_id)

    # Ensure non-admin users only edit awards from their department
    user_profile = get_object_or_404(UserProfile, user=request.user)
    if not request.user.is_superuser and award.department != user_profile.department:
        messages.error(request, "You don't have permission to edit this award.")
        return redirect('award_list')

    if request.method == "POST":
        award.innovation_title = request.POST.get('innovation_title', award.innovation_title)
        award.awardee_name = request.POST.get('awardee_name', award.awardee_name)
        award.awarding_agency = request.POST.get('awarding_agency', award.awarding_agency)
        award.award_year = request.POST.get('award_year', award.award_year)
        award.category = request.POST.get('category', award.category)

        # Handle document upload (if provided)
        if 'document' in request.FILES:
            award.document = request.FILES['document']

        award.save()
        messages.success(request, "Award updated successfully!")

        # Return JSON response for AJAX support
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Award updated successfully!'})

        return redirect('award_list')

    return redirect('award_list')

def award_delete(request, award_id):
    """Delete an award and show a message without redirecting."""
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)

    award = get_object_or_404(AwardRecognition, id=award_id)

    try:
        award.delete()
        messages.success(request, "Award deleted successfully!")
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@login_required
def add_award(request):
    user = request.user

    if request.method == "POST":
        if "excel_file" in request.FILES:  # Bulk Upload Handling
            excel_file = request.FILES["excel_file"]

            # Validate file format
            if not excel_file.name.endswith(".xlsx"):
                messages.error(request, "Invalid file format. Please upload an Excel file (.xlsx).")
                return redirect("add_award")

            try:
                df = pd.read_excel(excel_file).astype(str).fillna('')  # Convert all values to string & replace NaN

                # Required Columns
                base_required_columns = ["Innovation Title", "Awardee Name", "Awarding Agency", "Year", "Category"]
                required_columns = base_required_columns + ["Department"] if user.is_superuser else base_required_columns

                # Check for Missing Columns
                missing_cols = [col for col in required_columns if col not in df.columns]
                if missing_cols:
                    messages.error(request, f"Missing columns in Excel: {', '.join(missing_cols)}")
                    return redirect("add_award")

                skipped_rows = []  # Store rows with errors
                for index, row in df.iterrows():
                    department = None

                    # Assign Department
                    if user.is_superuser:
                        department_name = str(row.get("Department", "")).strip()
                        department = Department.objects.filter(name=department_name).first()
                    else:
                        department = get_object_or_404(UserProfile, user=user).department

                    if not department:
                        skipped_rows.append(f"Row {index+2}: Invalid Department ({row.get('Department', 'N/A')})")
                        continue

                    # Validate Award Year
                    try:
                        award_year = int(float(row["Year"]))  # Convert float values like 2022.0 to int
                    except (ValueError, TypeError):
                        skipped_rows.append(f"Row {index+2}: Invalid Year ({row['Year']})")
                        continue

                    # Create Award Entry (Document remains None in bulk upload)
                    AwardRecognition.objects.create(
                        department=department,
                        innovation_title=row["Innovation Title"].strip(),
                        awardee_name=row["Awardee Name"].strip(),
                        awarding_agency=row["Awarding Agency"].strip(),
                        award_year=award_year,
                        category=row["Category"].strip(),
                        document=None,  # Bulk uploads do not support document uploads
                    )

                # Show Skipped Rows
                if skipped_rows:
                    messages.warning(request, "Some rows were skipped due to errors:\n" + "\n".join(skipped_rows))

                messages.success(request, "Awards added successfully from Excel!")
                return redirect("award_list")

            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect("add_award")

        else:  # Single Entry Form Submission
            try:
                # Assign Department
                if user.is_superuser:
                    department_code = request.POST.get("department_code")
                    department = get_object_or_404(Department, department_code=department_code)
                else:
                    department = get_object_or_404(UserProfile, user=user).department

                # Get Form Data
                innovation_title = request.POST.get("innovation_title", "").strip()
                awardee_name = request.POST.get("awardee_name", "").strip()
                awarding_agency = request.POST.get("awarding_agency", "").strip()
                award_year = request.POST.get("award_year", "").strip()
                category = request.POST.get("category", "").strip()
                document = request.FILES.get("document")  # Upload Document

                # Validate Required Fields
                if not all([innovation_title, awardee_name, awarding_agency, award_year, category]):
                    messages.error(request, "All fields are required except document.")
                    return redirect("add_award")

                # Validate Year
                try:
                    award_year = int(award_year)
                except ValueError:
                    messages.error(request, "Invalid year format. Please enter a valid numeric year.")
                    return redirect("add_award")

                # Create Award Entry
                AwardRecognition.objects.create(
                    department=department,
                    innovation_title=innovation_title,
                    awardee_name=awardee_name,
                    awarding_agency=awarding_agency,
                    award_year=award_year,
                    category=category,
                    document=document,  # Save uploaded document
                )

                messages.success(request, "Award added successfully!")
                return redirect("award_list")

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect("add_award")

    # Get Departments for Superusers
    departments = Department.objects.all() if user.is_superuser else None
    return render(request, "AddData/add_award.html", {"departments": departments,'disable_filter': True})



@login_required
def patent_list(request):
    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')
    """View to display patents. Superusers see all patents, others see only their department's patents."""
    if request.user.is_superuser:
        patents = Patent.objects.all()
    else:
        patents = Patent.objects.filter(department=request.user.userprofile.department)

    return render(request, 'Forms/patent_list.html', {'patents': patents,'department':selected_department,'criterion_id':criterion_id})

@login_required
def add_patent(request):
    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')
    """Allow users to add patents (restricted to their department)."""
    if request.method == "POST":
        patenter_name = request.POST.get('patenter_name')
        patent_number = request.POST.get('patent_number')
        title = request.POST.get('patent_title')
        award_year = request.POST.get('award_year')
        document = request.FILES.get('document')

        # Assign department from UserProfile
        department = request.user.userprofile.department if not request.user.is_superuser else None

        patent = Patent.objects.create(
            patenter_name=patenter_name,
            patent_number=patent_number,
            title=title,
            award_year=award_year,
            document=document,
            department=department
        )
        messages.success(request, "Patent added successfully!")
        return redirect('patent_list')

    return render(request, 'AddData/add_patent.html',{'department':selected_department,'criterion_id':criterion_id,'disable_filter': True})

@login_required
def edit_patent(request, patent_id):
    """View to edit a patent. Restricts access based on department."""
    patent = get_object_or_404(Patent, id=patent_id)

    if not request.user.is_superuser and patent.department != request.user.userprofile.department:
        messages.error(request, "You do not have permission to edit this patent.")
        return redirect('patent_list')

    if request.method == "POST":
        patent.patenter_name = request.POST.get('patenter_name')
        patent.patent_number = request.POST.get('patent_number')
        patent.title = request.POST.get('title')
        patent.award_year = request.POST.get('award_year')

        if 'document' in request.FILES:
            patent.document = request.FILES['document']

        patent.save()
        messages.success(request, "Patent updated successfully!")
        return redirect('patent_list')

    return render(request, 'edit_patent.html', {'patent': patent})

@login_required
def delete_patent(request, patent_id):
    """View to delete a patent. Restricts access based on department."""
    patent = get_object_or_404(Patent, id=patent_id)

    if not request.user.is_superuser and patent.department != request.user.userprofile.department:
        messages.error(request, "You do not have permission to delete this patent.")
        return redirect('patent_list')

    patent.delete()
    messages.success(request, "Patent deleted successfully!")
    return redirect('patent_list')

@login_required
def phd_list(request):
    criterion_id = request.session.get('criterion_id')
    selected_department = request.GET.get('department', '')
    """Display Ph.D. records based on user role."""
    if request.user.is_superuser:
        phds = PhDAward.objects.all()
    else:
        phds = PhDAward.objects.filter(department=request.user.userprofile.department)
    
    return render(request, "Forms/phd_list.html", {"phds": phds,'criterion_id':criterion_id})

@login_required
def add_phd(request):
    """Add PhD Award - Restrict department for users."""
    departments = Department.objects.all()  # Get list of departments for admins

    if request.method == "POST":
        scholar_name = request.POST.get("scholar_name")
        guides = request.POST.get("guide")
        title = request.POST.get("thesis_title")
        registration_year = request.POST.get("registration_year")
        award_year = request.POST.get("award_year")
        document = request.FILES.get("document")

        if request.user.is_superuser:
            department_code = request.POST.get("department")  # Admin selects department
        else:
            department_code = request.user.userprofile.department.department_code  # Auto-assign for users

        # ✅ Convert department name to a Department instance
        department = get_object_or_404(Department, department_code=department_code)

        # ✅ Now save the PhDAward object
        PhDAward.objects.create(
            scholar_name=scholar_name,
            department=department,  # Assigning Department instance instead of string
            guides=guides,
            title=title,
            registration_year=registration_year,
            award_year=award_year,
            document=document
        )
        messages.success(request, "Ph.D. record added successfully.")
        return redirect("phd_list")

    return render(request, "AddData/phd_form.html", {"departments": departments})
@login_required
def edit_phd(request, phd_id):
    """Edit an existing Ph.D. record."""
    phd = get_object_or_404(PhDAward, id=phd_id)

    if request.method == "POST":
        # Retrieve updated data from the form
        phd.scholar_name = request.POST.get("scholar_name", phd.scholar_name)
        phd.guides = request.POST.get("guides", phd.guides)
        phd.title = request.POST.get("title", phd.title)
        phd.registration_year = request.POST.get("registration_year", phd.registration_year)
        phd.award_year = request.POST.get("award_year", phd.award_year)

        # Get the department code from the form
        department_code = request.POST.get("department", phd.department.department_code)

        # Retrieve the corresponding Department object
        phd.department = get_object_or_404(Department, department_code=department_code)

        # Save the updated record
        phd.save()
        messages.success(request, "Ph.D. record updated successfully.")
        return redirect("phd_list")

    # If not a POST request, redirect with an error message
    messages.error(request, "Failed to update the Ph.D. record.")
    return redirect("phd_list")
@login_required
def view_phd(request, phd_id):
    """Display details of a specific Ph.D. record."""
    phd = get_object_or_404(PhDAward, id=phd_id)
    return render(request, "viewData/phd_details.html", {"phd": phd})
@login_required
def delete_phd(request, phd_id):
    """Delete a Ph.D. record."""
    phd = get_object_or_404(PhDAward, id=phd_id)

    if request.method == "POST":  # Ensure this is a POST request
        phd.delete()
        messages.success(request, "Ph.D. record deleted successfully.")
        return redirect("phd_list")  # Redirect to the list view after success

    messages.error(request, "Failed to delete the record.")
    return redirect("phd_list")
@login_required
def download_phd_pdf(request, phd_id):
    """
    Generate and download a PDF for a specific Ph.D. record.
    """
    # Get the Ph.D. record
    phd = get_object_or_404(PhDAward, id=phd_id)
    

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="PhD_{phd.scholar_name}_Details.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # **Add Logo**
    try:
        logo_path = finders.find("images/logo2.png")  # Ensure the logo path is correct
        if logo_path:
            logo = ImageReader(logo_path)
            pdf.drawImage(logo, 50, height - 85, 140, 50, mask="auto")
        else:
            print("Logo file not found.")
    except Exception as e:
        print(f"Error loading logo: {e}")

    # **Header Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Ph.D. Scholar Details")

    # **Table Data**
    department_name = phd.department.department_name if phd.department else "N/A"
    data = [
        ["Scholar Name", phd.scholar_name],
        ["Department", department_name],
        ["Guide(s)", phd.guides],
        ["Title of the Thesis", phd.title],
        ["Year of Registration", phd.registration_year],
        ["Year of Award", phd.award_year],
        ["Document Available", "Yes" if phd.document else "No"],
    ]

    # **Style Table**
    table = Table(data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # **Draw Table**
    table_x = (width - 500) / 2
    table_y = height - 300
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
    pdf.drawCentredString(width / 2, 30, "© 2025 Your Organization Name. All Rights Reserved.")

    # **Curved Double Line Border**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response
@login_required
def view_research_papers(request):
    """
    View function to display research papers based on user restrictions.
    - Superusers: Can view all papers.
    - Regular users: Can only view their own papers and papers from their department.
    """
    if request.user.is_superuser:
        # Superusers can see all papers
        papers = ResearchPaper.objects.all()
    else:
        # Regular users can only see their papers and their department's papers
        user_department = request.user.userprofile.department  # Assuming UserProfile links users to departments
        papers = ResearchPaper.objects.filter(department=user_department)

    return render(request, "Forms/research_paper.html", {"papers": papers})
@login_required
def add_research_paper(request):
    """
    View function to handle the addition of a new research paper.
    """
    if request.method == "POST":
        # Get form data
        paper_title = request.POST.get("paper_title")
        author_name = request.POST.get("author_name")
        journal_name = request.POST.get("journal_name")
        publication_year = request.POST.get("publication_year")
        issn_number = request.POST.get("issn_number")
        ugc_link = request.POST.get("ugc_link")
        document = request.FILES.get("document")

        # Get the user's department
        user_department = request.user.userprofile.department

        # Save the research paper instance
        research_paper = ResearchPaper.objects.create(
            title=paper_title,
            authors=author_name,
            department=user_department,
            journal=journal_name,
            year=publication_year,
            issn=issn_number,
            ugc_link=ugc_link,
            document=document
        )
        research_paper.save()
        messages.success(request, "Research paper added successfully!")
        return redirect("view_research_papers")  # Redirect to list view

    # Render the form for GET requests
    return render(request, "AddData/add_research_paper.html")

@login_required
def edit_research_paper(request, paper_id):
    """
    Edit an existing Research Paper record.
    """
    research_paper = get_object_or_404(ResearchPaper, id=paper_id)

    if request.method == "POST":
        # Retrieve updated data from the form
        research_paper.title = request.POST.get("paper_title", research_paper.title)
        research_paper.authors = request.POST.get("author_name", research_paper.authors)
        research_paper.journal = request.POST.get("journal_name", research_paper.journal)
        research_paper.year = request.POST.get("publication_year", research_paper.year)
        research_paper.issn = request.POST.get("issn_number", research_paper.issn)
        research_paper.ugc_link = request.POST.get("ugc_link", research_paper.ugc_link)

        # Handle document upload if a new one is provided
        if request.FILES.get("document"):
            research_paper.document = request.FILES["document"]

        # Save the record
        research_paper.save()
        messages.success(request, "Research Paper updated successfully!")
        return redirect("view_research_papers")

    # If not a POST request, redirect with an error message
    messages.error(request, "Failed to update the Research Paper.")
    return redirect("view_research_papers")
@login_required
def delete_research_paper(request, paper_id):
    """
    Delete a Research Paper record.
    """
    research_paper = get_object_or_404(ResearchPaper, id=paper_id)

    if request.method == "POST":
        research_paper.delete()
        messages.success(request, "Research Paper deleted successfully!")
        return redirect("view_research_papers")

    messages.error(request, "Failed to delete the Research Paper.")
    return redirect("view_research_papers")
@login_required
def view_research_paper(request, paper_id):
    """
    View details of a specific Research Paper.
    """
    paper = get_object_or_404(ResearchPaper, id=paper_id)
    return render(request, "viewData/research_paper_detail.html", {"paper": paper})
@login_required
def download_research_paper_pdf(request, paper_id):
    """
    Generate and download a PDF for a specific Research Paper.
    """
    # Get the Research Paper record
    paper = get_object_or_404(ResearchPaper, id=paper_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ResearchPaper_{paper.title.replace(" ", "_")}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # **Add Logo**
    try:
        logo_path = finders.find("images/logo2.png")  # Ensure the logo path is correct
        if logo_path:
            logo = ImageReader(logo_path)
            pdf.drawImage(logo, 50, height - 85, 140, 50, mask="auto")
        else:
            print("Logo file not found.")
    except Exception as e:
        print(f"Error loading logo: {e}")

    # **Header Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Research Paper Details")

    # **Table Data**
    department_name = paper.department.department_name if paper.department else "N/A"
    data = [
        ["Title of Paper", paper.title],
        ["Authors", paper.authors],
        ["Department", department_name],
        ["Journal", paper.journal],
        ["Year of Publication", paper.year],
        ["ISSN", paper.issn],
        ["UGC Link", paper.ugc_link if paper.ugc_link else "No link available"],
    ]

    # **Style Table**
    table = Table(data, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # **Draw Table**
    table_x = (width - 500) / 2
    table_y = height - 300
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, f"Generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
    pdf.drawCentredString(width / 2, 30, "© 2025 Your Organization Name. All Rights Reserved.")

    # **Curved Double Line Border**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response

@login_required
def view_book_chapters(request):
    """
    View book/chapter records:
    - Teachers and scholars can only see their own records.
    - Department staff can see all records related to their department.
    """
    if request.user.userprofile.is_department_staff:
        # Department staff can view all records related to their department
        user_department = request.user.userprofile.department
        books = BookChapter.objects.filter(teacher_name__user_profile__department=user_department)
    else:
        # Other users (teachers, scholars) can only view their own records
        user_profile = request.user.userprofile
        books = BookChapter.objects.filter(teacher_name__user_profile=user_profile)

    return render(request, "Forms/book_chapter_list.html", {"books": books})
@login_required
def edit_book_chapter(request, book_id):
    """
    Edit a book/chapter record with access control.
    """
    book = get_object_or_404(BookChapter, id=book_id)

    # Restrict access based on user role
    if not request.user.userprofile.is_department_staff:
        # For non-department staff, restrict editing to their own data
        if book.teacher_name.user_profile != request.user.userprofile:
            messages.error(request, "You are not authorized to edit this record.")
            return redirect("view_book_chapters")

    if request.method == "POST":
        book.book_title = request.POST.get("book_title", book.book_title)
        book.paper_title = request.POST.get("paper_title", book.paper_title)
        book.proceedings_title = request.POST.get("proceedings_title", book.proceedings_title)
        book.conference_name = request.POST.get("conference_name", book.conference_name)
        book.national_international = request.POST.get("national_international", book.national_international)
        book.publication_year = request.POST.get("publication_year", book.publication_year)
        book.isbn_issn = request.POST.get("isbn_issn", book.isbn_issn)
        book.affiliating_institute = request.POST.get("affiliating_institute", book.affiliating_institute)
        book.publisher = request.POST.get("publisher", book.publisher)
        book.save()

        messages.success(request, "Book/Chapter record updated successfully!")
        return redirect("view_book_chapters")

    teachers = Teacher.objects.all()
    return render(request, "Forms/edit_book_chapter.html", {"book": book, "teachers": teachers})
@login_required
def delete_book_chapter(request, book_id):
    """
    Delete a book/chapter record with access control.
    """
    book = get_object_or_404(BookChapter, id=book_id)

    # Restrict access based on user role
    if not request.user.userprofile.is_department_staff:
        if book.teacher_name.user_profile != request.user.userprofile:
            messages.error(request, "You are not authorized to delete this record.")
            return redirect("view_book_chapters")

    if request.method == "POST":
        book.delete()
        messages.success(request, "Book/Chapter record deleted successfully!")
        return redirect("view_book_chapters")

    messages.error(request, "Failed to delete the record.")
    return redirect("view_book_chapters")
@login_required
def add_book_chapter(request):
    """
    Add a new book/chapter record.
    """
    if request.method == "POST":
        try:
            # Retrieve data from the POST request
            teacher_id = request.POST.get("teacher_name")
            book_title = request.POST.get("book_title")
            paper_title = request.POST.get("paper_title")
            proceedings_title = request.POST.get("proceedings_title")
            conference_name = request.POST.get("conference_name")
            national_international = request.POST.get("conference_type")  # Map to dropdown name
            publication_year = request.POST.get("publication_year")
            isbn_issn = request.POST.get("isbn")
            affiliating_institute = request.POST.get("affiliating_institute")
            publisher = request.POST.get("publisher")

            # Validate National/International Dropdown
            if not national_international:
                messages.error(request, "Please select whether the conference is National or International.")
                return redirect("add_book_chapter")

            # Retrieve the teacher instance
            teacher = Teacher.objects.get(id=teacher_id)

            # Create the BookChapter record
            BookChapter.objects.create(
                teacher_name=teacher,
                book_title=book_title,
                paper_title=paper_title,
                proceedings_title=proceedings_title,
                conference_name=conference_name,
                national_international=national_international,
                publication_year=publication_year,
                isbn_issn=isbn_issn,
                affiliating_institute=affiliating_institute,
                publisher=publisher,
            )

            messages.success(request, "Book/Chapter record added successfully!")
            return redirect("view_book_chapters")
        except Teacher.DoesNotExist:
            messages.error(request, "Invalid teacher selected. Please choose a valid teacher.")
            return redirect("add_book_chapter")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect("add_book_chapter")

    # Render the form
    teachers = Teacher.objects.all()
    return render(request, "AddData/add_book_chapter.html", {"teachers": teachers})
@login_required
def view_book_chapter_details(request, book_id):
    """
    View details of a specific book/chapter record.
    """
    book = get_object_or_404(BookChapter, id=book_id)

    return render(request, "viewData/book_chapter_details.html", {"book": book})
@login_required
def download_book_chapter_pdf(request, book_id):
    """
    Generate and download a PDF for a specific Book/Chapter.
    """
    book = get_object_or_404(BookChapter, id=book_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="book_chapter_{book.id}.pdf"'

    # Create the PDF
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # ✅ Find Logo File
    logo_path = finders.find("images/logo2.png")
    if logo_path:
        logo = ImageReader(logo_path)
        pdf.drawImage(logo, 50, height - 85, width=140, height=50, mask="auto")

    # **Header Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Book/Chapter Details")

    # **Table Data**
    data = [
        ["Teacher Name", book.teacher_name.user_profile.user.username],
        ["Book/Chapter Title", book.book_title],
        ["Paper Title", book.paper_title or "N/A"],
        ["Proceedings Title", book.proceedings_title or "N/A"],
        ["Conference Name", book.conference_name or "N/A"],
        ["National/International", book.national_international],
        ["Year of Publication", book.publication_year],
        ["ISBN/ISSN", book.isbn_issn or "N/A"],
        ["Affiliating Institute", book.affiliating_institute or "N/A"],
        ["Publisher", book.publisher],
    ]

    # Table Style
    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # **Draw Table in the Center**
    table_x = (width - 480) / 2
    table_y = height - 280
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # **Footer Section**
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Page Borders**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response

@login_required
def demand_ratio_list(request):
    """
    Display a list of Demand Ratio records with department-level restrictions.
    Superusers can view all records, while department users see only their department's data.
    """
    # Check if the user is a superuser
    if request.user.is_superuser:
        # Superusers can view all data
        records = DemandRatio.objects.all()
    else:
        # Department users can only view their department's data
        if hasattr(request.user, "userprofile") and hasattr(request.user.userprofile, "department"):
            user_department = request.user.userprofile.department
            records = DemandRatio.objects.filter(department=user_department)
        else:
            # If the user has no associated department, return an empty queryset
            records = DemandRatio.objects.none()

    # Apply filters if provided
    programme_name = request.GET.get('programme_name', '').strip()
    programme_code = request.GET.get('programme_code', '').strip()
    year_from = request.GET.get('year_from', '').strip()
    year_to = request.GET.get('year_to', '').strip()

    if programme_name:
        records = records.filter(programme_name__icontains=programme_name)

    if programme_code:
        records = records.filter(programme_code__icontains=programme_code)

    if year_from and year_to:
        records = records.filter(academic_year__gte=year_from, academic_year__lte=year_to)
    elif year_from:
        records = records.filter(academic_year__gte=year_from)
    elif year_to:
        records = records.filter(academic_year__lte=year_to)

    context = {
        'records': records,
    }

    return render(request, 'Forms/demand_ratio_list.html', context)
@login_required
def download_demand_ratio_pdf(request, record_id):
    """
    Generate and download a PDF for a specific Demand Ratio record.
    """
    record = get_object_or_404(DemandRatio, id=record_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="demand_ratio_{record.id}.pdf"'

    # Create the PDF
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # Header Section with Title
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Demand Ratio Details")

    # Table Data
    data = [
        ["Programme Name", record.programme_name],
        ["Programme Code", record.programme_code],
        ["Seats Available", record.num_seats],
        ["Eligible Applications", record.num_applications],
        ["Students Admitted", record.num_students_admitted],
        ["Academic Year", record.academic_year],
    ]

    # Style the Table
    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # Draw Table in the Center
    table_x = (width - 480) / 2
    table_y = height - 300
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # Footer Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Page Borders**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response

@login_required
def add_demand_ratio(request):
    if request.method == "POST":
        academic_year = request.POST.get("academic_year")
        programme_name = request.POST.get("programme_name")
        programme_code = request.POST.get("programme_code")
        num_seats = request.POST.get("num_seats")
        num_applications = request.POST.get("num_applications")
        num_students_admitted = request.POST.get("num_students_admitted")
        department_code = request.POST.get("department")

        # Convert numeric fields
        try:
            num_seats = int(num_seats)
            num_applications = int(num_applications)
            num_students_admitted = int(num_students_admitted)
        except ValueError:
            return render(request, "AddData/add_demand_ratio.html", {"error": "Invalid input values!"})

        # Restrict department selection for department users
        if not request.user.is_superuser:
            department = request.user.userprofile.department
        else:
            department = Department.objects.get(department_code=department_code)

        # Save the data
        DemandRatio.objects.create(
            academic_year=academic_year,
            programme_name=programme_name,
            programme_code=programme_code,
            num_seats=num_seats,
            num_applications=num_applications,
            num_students_admitted=num_students_admitted,
            department=department
        )

        return redirect("demand_ratio_list")  # Redirect to the list page after saving

    # Generate academic year options in yyyy-yyyy format
    from datetime import datetime
    current_year = datetime.now().year
    academic_years = [f"{y}-{y+1}" for y in range(2000, current_year + 1)]

    departments = Department.objects.all() if request.user.is_superuser else [request.user.userprofile.department]

    return render(request, "AddData/add_demand_ratio.html", {"academic_years": academic_years, "departments": departments})
@login_required
def edit_demand_ratio(request, record_id):
    """
    Edit an existing Demand Ratio record.
    """
    record = get_object_or_404(DemandRatio, id=record_id)

    if request.method == "POST":
        record.programme_name = request.POST.get("programme_name", record.programme_name)
        record.programme_code = request.POST.get("programme_code", record.programme_code)
        record.num_seats = request.POST.get("num_seats", record.num_seats)
        record.num_applications = request.POST.get("num_applications", record.num_applications)
        record.num_students_admitted = request.POST.get("num_students_admitted", record.num_students_admitted)
        record.academic_year = request.POST.get("academic_year", record.academic_year)

        try:
            record.save()
            messages.success(request, "Demand Ratio record updated successfully!")
        except Exception as e:
            messages.error(request, f"Error updating record: {e}")

        return redirect("demand_ratio_list")

    messages.error(request, "Invalid request method.")
    return redirect("demand_ratio_list")
@login_required
def view_demand_ratio_details(request, record_id):
    """
    View details of a specific Demand Ratio record.
    """
    record = get_object_or_404(DemandRatio, id=record_id)
    return render(request, "viewData/demand_ratio_details.html", {"record": record})
@login_required
def delete_demand_ratio(request, record_id):
    """
    Delete a specific Demand Ratio record.
    """
    record = get_object_or_404(DemandRatio, id=record_id)

    if request.method == "POST":
        try:
            record.delete()
            messages.success(request, "Demand Ratio record deleted successfully!")
        except Exception as e:
            messages.error(request, f"Error deleting record: {e}")
        return redirect("demand_ratio_list")

    messages.error(request, "Invalid request method.")
    return redirect("demand_ratio_list")

def admitted_student_list(request):
    students = AdmittedStudent.objects.select_related('department').all()
    return render(request,'Forms/admited_students.html',{'records': students})
@login_required
def add_admitted_student(request):
    """Handle form submission for adding a new admitted student"""
    
    # Get the logged-in user's department
    try:
        user_department = request.user.userprofile.department
    except AttributeError:
        user_department = None

    if request.method == "POST":
        year = request.POST.get('year', '')
        programme_name = request.POST.get('programme_name', '')
        department_id = request.POST.get('department', '')

        sc_earmarked = request.POST.get('sc_earmarked', 0)
        st_earmarked = request.POST.get('st_earmarked', 0)
        obc_earmarked = request.POST.get('obc_earmarked', 0)
        gen_earmarked = request.POST.get('gen_earmarked', 0)
        others_earmarked = request.POST.get('others_earmarked', 0)

        sc_admitted = request.POST.get('sc_admitted', 0)
        st_admitted = request.POST.get('st_admitted', 0)
        obc_admitted = request.POST.get('obc_admitted', 0)
        gen_admitted = request.POST.get('gen_admitted', 0)
        others_admitted = request.POST.get('others_admitted', 0)

        # Ensure the department belongs to the user
        if user_department and str(user_department.department_code) == department_id:
            department = user_department
        else:
            return redirect('add_admitted_student')

        # Create AdmittedStudent entry
        AdmittedStudent.objects.create(
            year=year,
            programme_name=programme_name,
            department=department,
            sc_earmarked=sc_earmarked,
            st_earmarked=st_earmarked,
            obc_earmarked=obc_earmarked,
            gen_earmarked=gen_earmarked,
            others_earmarked=others_earmarked,
            sc_admitted=sc_admitted,
            st_admitted=st_admitted,
            obc_admitted=obc_admitted,
            gen_admitted=gen_admitted,
            others_admitted=others_admitted,
        )
        return redirect('admitted_students_list')

    # Only pass the logged-in user's department
    departments = Department.objects.filter(department_code=user_department.department_code) if user_department else []

    return render(request, 'AddData/add_admitted_student.html', {'departments': departments})

@login_required
def edit_admitted_student(request, record_id):
    """
    Edit an existing admitted student record.
    """
    record = get_object_or_404(AdmittedStudent, id=record_id)

    if request.method == "POST":
        try:
            record.programme_name = request.POST.get("programme_name", record.programme_name)
            record.year = request.POST.get("year", record.year)
            record.sc_earmarked = request.POST.get("sc_earmarked", record.sc_earmarked)
            record.st_earmarked = request.POST.get("st_earmarked", record.st_earmarked)
            record.obc_earmarked = request.POST.get("obc_earmarked", record.obc_earmarked)
            record.gen_earmarked = request.POST.get("gen_earmarked", record.gen_earmarked)
            record.others_earmarked = request.POST.get("others_earmarked", record.others_earmarked)
            record.sc_admitted = request.POST.get("sc_admitted", record.sc_admitted)
            record.st_admitted = request.POST.get("st_admitted", record.st_admitted)
            record.obc_admitted = request.POST.get("obc_admitted", record.obc_admitted)
            record.gen_admitted = request.POST.get("gen_admitted", record.gen_admitted)
            record.others_admitted = request.POST.get("others_admitted", record.others_admitted)

            # Save updated record
            record.save()
            messages.success(request, "Admitted student record updated successfully!")
        except Exception as e:
            messages.error(request, f"Error updating record: {str(e)}")

        return redirect("admitted_students_list")

    messages.error(request, "Invalid request method.")
    return redirect("admitted_students_list")
@login_required
def delete_admitted_student(request, record_id):
    """
    Delete an admitted student record.
    """
    record = get_object_or_404(AdmittedStudent, id=record_id)

    if request.method == "POST":
        try:
            record.delete()
            messages.success(request, "Admitted student record deleted successfully!")
        except Exception as e:
            messages.error(request, f"Error deleting record: {str(e)}")
        return redirect("admitted_students_list")

    messages.error(request, "Invalid request method.")
    return redirect("admitted_students_list")
@login_required
def view_admitted_student_details(request, record_id):
    """
    View details of a specific Admitted Student.
    """
    record = get_object_or_404(AdmittedStudent, id=record_id)
    return render(request, "viewData/admitted_student_details.html", {"record": record})
@login_required
def download_admitted_student_pdf(request, record_id):
    """
    Generate and download a PDF for a specific Admitted Student record.
    """
    # Get the specific Admitted Student record
    record = get_object_or_404(AdmittedStudent, id=record_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="admitted_student_{record.id}.pdf"'

    # Create the PDF
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # Header Section with Title
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Admitted Student Details")

    # Table Data
    data = [
        ["Year", record.year],
        ["Programme Name", record.programme_name],
        ["SC (Earmarked)", record.sc_earmarked],
        ["ST (Earmarked)", record.st_earmarked],
        ["OBC (Earmarked)", record.obc_earmarked],
        ["General (Earmarked)", record.gen_earmarked],
        ["Others (Earmarked)", record.others_earmarked],
        ["SC (Admitted)", record.sc_admitted],
        ["ST (Admitted)", record.st_admitted],
        ["OBC (Admitted)", record.obc_admitted],
        ["General (Admitted)", record.gen_admitted],
        ["Others (Admitted)", record.others_admitted],
    ]

    # Style the Table
    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # Draw Table in the Center
    table_x = (width - 480) / 2
    table_y = height - 300
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # Footer Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Page Borders**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response
@login_required
def teacher_serving_post_list(request):
    """
    View to list all Teacher Serving Post records with optional filters.
    """
    # Retrieve all records with related teacher information
    records = TeacherServingPost.objects.select_related('teacher').all()

    # Filters (based on POST data from filter form)
    teacher_name = request.POST.get('teacher_name', '').strip()
    department = request.POST.get('department', '').strip()
    designation = request.POST.get('designation', '').strip()

    if teacher_name:
        records = records.filter(teacher__user_profile__user__username__icontains=teacher_name)
    if department:
        records = records.filter(teacher__department__icontains=department)
    if designation:
        records = records.filter(teacher__position__icontains=designation)

    # Render the list template with filtered records
    return render(request, 'Forms/teacher_serving_post_list.html', {'records': records})
@login_required
def add_teacher_serving_post(request):
    user_profile = get_object_or_404(UserProfile, user=request.user)

    if user_profile.is_teacher:
        try:
            teacher = Teacher.objects.get(user_profile=user_profile)
        except Teacher.DoesNotExist:
            messages.error(request, "You are not associated with a teacher record.")
            return redirect("teacher_serving_post_list")

        if request.method == "POST":
            # Retrieve the last_year_service value from the POST request
            last_year_service = request.POST.get("last_year_service", "").strip() or None
            appointment_year = request.POST.get("appointment_year").strip()
            nature_of_appointment = request.POST.get("nature_of_appointment").strip()
            experience_years = request.POST.get("experience_years").strip()
            is_serving = request.POST.get("is_serving") == "Yes"

            try:
                # Save to the TeacherServingPost model
                TeacherServingPost.objects.create(
                    teacher=teacher,
                    appointment_year=appointment_year,
                    nature_of_appointment=nature_of_appointment,
                    experience_years=experience_years,
                    is_serving=is_serving,
                    last_year_service=last_year_service
                )
                messages.success(request, "Teacher Serving Post record added successfully.")
                return redirect("teacher_serving_post_list")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

        return render(request, 'AddData/add_teacher_serving_post.html', {"is_staff": False})

    elif user_profile.is_department_staff:
        teachers = Teacher.objects.filter(user_profile__department=user_profile.department)

        if request.method == "POST":
            teacher_id = request.POST.get("teacher")
            last_year_service = request.POST.get("last_year_service", "").strip() or None
            appointment_year = request.POST.get("appointment_year").strip()
            nature_of_appointment = request.POST.get("nature_of_appointment").strip()
            experience_years = request.POST.get("experience_years").strip()
            is_serving = request.POST.get("is_serving") == "Yes"

            try:
                teacher = get_object_or_404(Teacher, id=teacher_id)
                TeacherServingPost.objects.create(
                    teacher=teacher,
                    appointment_year=appointment_year,
                    nature_of_appointment=nature_of_appointment,
                    experience_years=experience_years,
                    is_serving=is_serving,
                    last_year_service=last_year_service
                )
                messages.success(request, "Teacher Serving Post record added successfully.")
                return redirect("teacher_serving_post_list")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

        return render(request, 'AddData/add_teacher_serving_post.html', {"teachers": teachers, "is_staff": True})

    else:
        messages.error(request, "You do not have permission to access this page.")
        return redirect("teacher_serving_post_list")

@login_required
def edit_teacher_serving_post(request, record_id):
    """
    View to edit a specific Teacher Serving Post record and route to the same Teacher Serving Post list page.
    """
    teacher_serving_post = get_object_or_404(TeacherServingPost, id=record_id)

    if request.method == "POST":
        appointment_year = request.POST.get("appointment_year", "").strip()
        nature_of_appointment = request.POST.get("nature_of_appointment", "").strip()
        experience_years = request.POST.get("experience_years", "").strip()
        is_serving = request.POST.get("is_serving") == "Yes"
        last_year_service = request.POST.get("last_year_service", "").strip() or None

        try:
            # Update the record
            teacher_serving_post.appointment_year = appointment_year
            teacher_serving_post.nature_of_appointment = nature_of_appointment
            teacher_serving_post.experience_years = experience_years
            teacher_serving_post.is_serving = is_serving
            teacher_serving_post.last_year_service = last_year_service
            teacher_serving_post.save()

            # Success message
            messages.success(request, "Record updated successfully.")
        except Exception as e:
            # Error message
            messages.error(request, f"An error occurred while updating: {e}")

        # Redirect to the same page (Teacher Serving Post list)
        return redirect("teacher_serving_post_list")

    # For GET requests, redirect to the Teacher Serving Post list (route to the same page)
    return redirect("teacher_serving_post_list")
@login_required
def delete_teacher_serving_post(request, record_id):
    """
    View to delete a specific Teacher Serving Post record and route to the same list page with messages.
    """
    teacher_serving_post = get_object_or_404(TeacherServingPost, id=record_id)

    if request.method == "POST":
        try:
            # Delete the record
            teacher_serving_post.delete()
            # Success message
            messages.success(request, "Record deleted successfully.")
        except Exception as e:
            # Error message
            messages.error(request, f"An error occurred while deleting: {e}")

        # Redirect to the same page (list page)
        return redirect("teacher_serving_post_list")

    # If GET request, redirect to the list page
    return redirect("teacher_serving_post_list")
@login_required
def full_time_teacher_list(request):
    """
    View to display Full-Time Teacher records.
    - If the user is a teacher, show their records only.
    - If the user is department staff, show all records related to their department.
    """
    user_profile = get_object_or_404(UserProfile, user=request.user)

    # Initialize an empty queryset
    records = FullTimeTeacher.objects.none()

    if user_profile.is_teacher:
        # Get the teacher object related to the logged-in user
        teacher = get_object_or_404(Teacher, user_profile=user_profile)
        # Fetch only records related to this teacher
        records = FullTimeTeacher.objects.filter(teacher_name=teacher)

    elif user_profile.is_department_staff:
        # Fetch all teachers related to the department of the logged-in staff member
        teachers_in_department = Teacher.objects.filter(user_profile__department=user_profile.department)
        # Fetch all FullTimeTeacher records related to those teachers
        records = FullTimeTeacher.objects.filter(teacher_name__in=teachers_in_department)

    # Pass the filtered records to the template
    return render(request, "Forms/teacher.html", {"records": records})
def add_teacher(request):
    """
    View to handle adding a new Full-Time Teacher record based on user roles.
    """
    user_profile = request.user.userprofile  # Get the user's profile to determine role

    if request.method == "POST":
        # If department staff, get the selected teacher from the dropdown; otherwise, get the logged-in teacher
        teacher_name_id = (
            request.POST.get("teacher") if user_profile.is_department_staff
            else Teacher.objects.get(user_profile=user_profile).id
        )
        qualification_year = request.POST.get("qualification")
        is_research_guide = request.POST.get("is_research_guide") == "Yes"
        recognition_year = request.POST.get("recognition_year")

        try:
            # Fetch the teacher object
            teacher_name = get_object_or_404(Teacher, id=teacher_name_id)

            # Save the new record in the FullTimeTeacher table
            FullTimeTeacher.objects.create(
                teacher_name=teacher_name,
                qualification_year=qualification_year,
                is_research_guide=is_research_guide,
                year_of_recognition=recognition_year,
            )

            # Success message
            messages.success(request, "Teacher record added successfully.")
            return redirect("full_time_teacher_list")  # Redirect to the list page
        except Exception as e:
            # Error message
            messages.error(request, f"An error occurred: {str(e)}")

    # Prepare context for GET requests
    if user_profile.is_department_staff:
        teachers = Teacher.objects.filter(user_profile__department=user_profile.department)
    else:
        teachers = None  # No dropdown for individual teachers

    return render(
        request,
        "AddData/add_teacher.html",
        {
            "teachers": teachers,
            "is_department_staff": user_profile.is_department_staff,
            "user": request.user,
        },
    )
@login_required
def edit_teacher(request, record_id):
    """
    View to edit a specific Full-Time Teacher record and route to the same Teacher List page.
    """
    teacher_record = get_object_or_404(FullTimeTeacher, id=record_id)

    if request.method == "POST":
        # Fetch form data
        qualification_year = request.POST.get("qualification_year", "").strip()
        is_research_guide = request.POST.get("is_research_guide") == "Yes"
        recognition_year = request.POST.get("year_of_recognition", "").strip()

        # Handle empty recognition year
        recognition_year = int(recognition_year) if recognition_year else None

        try:
            # Update the record
            teacher_record.qualification_year = qualification_year
            teacher_record.is_research_guide = is_research_guide
            teacher_record.year_of_recognition = recognition_year
            teacher_record.save()

            # Success message
            messages.success(request, "Record updated successfully.")
        except Exception as e:
            # Error message
            messages.error(request, f"An error occurred while updating: {e}")

        # Redirect to the same page (Teacher List page)
        return redirect("full_time_teacher_list")

    # Redirect to the Teacher List page for GET requests
    return redirect("full_time_teacher_list")

@login_required
def delete_teacher(request, record_id):
    """
    View to delete a specific Full-Time Teacher record and route to the same Teacher List page.
    """
    teacher_record = get_object_or_404(FullTimeTeacher, id=record_id)

    if request.method == "POST":
        try:
            # Delete the record
            teacher_record.delete()

            # Success message
            messages.success(request, "Record deleted successfully.")
        except Exception as e:
            # Error message
            messages.error(request, f"An error occurred while deleting: {str(e)}")

        # Redirect to the teacher list page
        return redirect("full_time_teacher_list")

    # Redirect to the teacher list page for non-POST requests
    return redirect("full_time_teacher_list")
@login_required
def teacher_sanctioned_post_list(request):
    """
    View to display teacher records.
    - If the user is a teacher, display only their records.
    - If the user is department staff, display all records related to their department.
    """
    user_profile = get_object_or_404(UserProfile, user=request.user)

    # Initialize queryset
    teachers = TeacherAgainstSanctionedPost.objects.none()

    if user_profile.is_teacher:
        # For logged-in teacher, fetch only their data
        teacher = get_object_or_404(Teacher, user_profile=user_profile)
        teachers = TeacherAgainstSanctionedPost.objects.filter(teacher=teacher)

    elif user_profile.is_department_staff:
        # For department staff, fetch all teachers in their department
        department_teachers = Teacher.objects.filter(user_profile__department=user_profile.department)
        teachers = TeacherAgainstSanctionedPost.objects.filter(teacher__in=department_teachers)

    # Render the template with filtered records
    return render(request, "Forms/teacher_sanctioned_post_list.html", {"teachers": teachers})

@login_required
def add_against_sanctioned_post(request):
    """
    View to handle adding a new Teacher Against Sanctioned Post record.
    """
    user_profile = get_object_or_404(UserProfile, user=request.user)
    teacher_data = None
    department_teachers = None

    # Role-based logic
    if user_profile.is_teacher:
        # Fetch data for the logged-in teacher
        teacher = get_object_or_404(Teacher, user_profile=user_profile)
        teacher_data = {
            "name": teacher.name,
            "pan": teacher.pan,
            "designation": teacher.position,
            "department": teacher.user_profile.department.department_name,
        }
    elif user_profile.is_department_staff:
        # Fetch all teachers belonging to the department
        department_teachers = Teacher.objects.filter(user_profile__department=user_profile.department)

    if request.method == "POST":
        if user_profile.is_department_staff:
            # Get the selected teacher from the dropdown
            teacher_id = request.POST.get("teacher_id")
            teacher = get_object_or_404(Teacher, id=teacher_id)
        else:
            # Fetch teacher for logged-in user
            teacher = get_object_or_404(Teacher, user_profile=user_profile)

        # Collect form data
        year_of_appointment = request.POST.get("yearOfAppointment", "").strip()
        nature_of_appointment = request.POST.get("natureOfAppointment", "").strip()
        experience_years = request.POST.get("experience", "").strip()
        is_serving = request.POST.get("isServing", "").strip() == "Yes"
        last_year_of_service = request.POST.get("lastYearOfService", "").strip()
        last_year_of_service = int(last_year_of_service) if not is_serving and last_year_of_service else None

        try:
            # Save the record
            TeacherAgainstSanctionedPost.objects.create(
                teacher=teacher,
                year_of_appointment=year_of_appointment,
                nature_of_appointment=nature_of_appointment,
                years_of_experience=experience_years,
                still_serving="Yes" if is_serving else "No",
                last_year_of_service=last_year_of_service,
            )
            messages.success(request, "Record added successfully.")
            return redirect("teacher_sanctioned_post_list")
        except Exception as e:
            messages.error(request, f"Error while adding record: {e}")

    return render(
        request,
        "AddData/add_against_sanctioned_post.html",
        {
            "is_teacher": user_profile.is_teacher,
            "teacher_data": teacher_data,
            "department_teachers": department_teachers,
        },
    )

@login_required
def edit_teacher_against_sanctioned_post(request, post_id):
    """
    View to edit a specific Teacher Against Sanctioned Post record and route to the same list page.
    """
    post_record = get_object_or_404(TeacherAgainstSanctionedPost, id=post_id)

    if request.method == "POST":
        # Fetch and validate form data
        year_of_appointment = request.POST.get("yearOfAppointment", "").strip()
        nature_of_appointment = request.POST.get("natureOfAppointment", "").strip()
        experience_years = request.POST.get("experience", "").strip()
        is_serving = request.POST.get("isServing", "").strip() == "Yes"
        last_year_of_service = request.POST.get("lastYearOfService", "").strip()

        try:
            # Validate required fields
            if not year_of_appointment or not year_of_appointment.isdigit():
                raise ValueError("Year of Appointment must be a valid number.")

            if not experience_years or not experience_years.isdigit():
                raise ValueError("Years of Experience must be a valid number.")

            # Handle empty "Last Year of Service"
            last_year_of_service = int(last_year_of_service) if not is_serving and last_year_of_service else None

            # Update the record
            post_record.year_of_appointment = int(year_of_appointment)
            post_record.nature_of_appointment = nature_of_appointment
            post_record.years_of_experience = int(experience_years)
            post_record.still_serving = "Yes" if is_serving else "No"
            post_record.last_year_of_service = last_year_of_service
            post_record.save()

            # Success message
            messages.success(request, "Record updated successfully.")
        except ValueError as ve:
            # Handle validation errors
            messages.error(request, str(ve))
        except Exception as e:
            # Generic error handler
            messages.error(request, f"An error occurred while updating: {e}")

        # Redirect back to the list page
        return redirect("teacher_sanctioned_post_list")

    # Redirect to the list page for GET requests
    return redirect("teacher_sanctioned_post_list")

@login_required
def delete_teacher_against_sanctioned_post(request, post_id):
    """
    View to delete a specific Teacher Against Sanctioned Post record.
    """
    post = get_object_or_404(TeacherAgainstSanctionedPost, id=post_id)

    try:
        post.delete()
        messages.success(request, "Record deleted successfully.")
    except Exception as e:
        messages.error(request, f"An error occurred while deleting: {e}")

    # Redirect to the same list page after deletion
    return redirect("teacher_sanctioned_post_list")

def e_governance(request):
    return render(request,"Forms/e_governance.html")
def add_e_governance(request):
    return render(request,"AddData/add_e_governance.html")
def conference(request):
    return render(request,"Forms/conference.html")
def add_conference(request):
    return render(request,"AddData/add_conference.html")
def training_record(request):
    return render(request,"Forms/training_record.html")
def add_training_record(request):
    return render(request,"AddData/add_training_record.html")
def faculty_development_program(request):
    return render(request,"Forms/faculty_development_program.html")
def add_faculty_development_program(request):
    return render(request,"AddData/add_faculty_development_program.html")
def grant_record(request):
    return render(request,"Forms/grant.html")
def add_grant_record(request):
    return render(request,"AddData/add_grand.html")
def quality_assurance(request):
    return render(request,"Forms/quality_assurance.html")
def add_quality_assurance(request): 
    return render(request,"AddData/add_quality_assurance.html")
def scholarship_list(request):
    return render(request,"Forms/scholarship.html")
def add_scholarship(request):
    return render(request,"AddData/add_scholarship.html")
def career_counseling(request):
    return render(request,"Forms/career_counseling.html")
def add_career_counseling(request):
    return render(request,"AddData/add_career_counseling.html")
from django.utils.timezone import now
from django.db.models import Q
from datetime import datetime
@login_required
def programme_list(request):
    user_profile = request.user.userprofile

    # Base query
    if request.user.is_superuser:
        programmes = Programme.objects.all()
    elif user_profile.is_department_staff:
        programmes = Programme.objects.filter(department=user_profile.department)
    else:
        programmes = Programme.objects.none()  # No access

    # Get filter parameters
    from_year = request.GET.get("fromYear")
    to_year = request.GET.get("toYear")
    cbcs_status = request.GET.get("programmeFilter")
    search_query = request.GET.get("searchBar", "").strip()

    # Construct filter query
    filter_conditions = Q()
    if from_year and from_year.isdigit():
        filter_conditions &= Q(year_of_introduction__gte=int(from_year))
    if to_year and to_year.isdigit():
        filter_conditions &= Q(year_of_introduction__lte=int(to_year))
    if cbcs_status:
        filter_conditions &= Q(cbcs_status=cbcs_status)
    if search_query:
        filter_conditions &= Q(programme_name__icontains=search_query)

    # Apply filters
    programmes = programmes.filter(filter_conditions)

    # Year range for dropdowns
    current_year = now().year
    year_range = [year for year in range(1999, current_year + 1)]

    # Pass context
    context = {
        "programmes": programmes,
        "year_range": year_range,
        "from_year": from_year,
        "to_year": to_year,
        "cbcs_status": cbcs_status,
        "search_query": search_query,
    }
    return render(request, "Forms/programme_list.html", context)


def add_programme(request):
    """
    View to handle the Add Programme form using POST.get() without using forms.
    """
    if request.method == "POST":
        # Fetch data directly from POST
        programme_code = request.POST.get("programme_code", "").strip()
        programme_name = request.POST.get("programme_name", "").strip()
        year_of_introduction = request.POST.get("year_of_introduction", "").strip()
        cbcs_status = request.POST.get("cbcs_status", "").strip()
        year_of_cbcs_implementation = request.POST.get("year_of_cbcs_implementation", "").strip()
        year_of_revision = request.POST.get("year_of_revision", "").strip()
        content_update_percentage = request.POST.get("content_update_percentage", "").strip()
        document_link = request.POST.get("document_link", "").strip()

        # Validate required fields
        if not programme_code or not programme_name or not year_of_introduction or not cbcs_status:
            messages.error(request, "Please fill out all required fields.")
            return render(request, "add_programme.html")

        # Convert numeric fields
        try:
            year_of_introduction = int(year_of_introduction)
            year_of_cbcs_implementation = int(year_of_cbcs_implementation) if year_of_cbcs_implementation else None
            year_of_revision = int(year_of_revision) if year_of_revision else None
            content_update_percentage = float(content_update_percentage) if content_update_percentage else None
        except ValueError:
            messages.error(request, "Please enter valid numeric values where applicable.")
            return render(request, "add_programme.html")

        # Create and save the Programme object
        programme = Programme(
            programme_code=programme_code,
            programme_name=programme_name,
            year_of_introduction=year_of_introduction,
            cbcs_status=cbcs_status,
            year_of_cbcs_implementation=year_of_cbcs_implementation,
            year_of_revision=year_of_revision,
            content_update_percentage=content_update_percentage,
            document_link=document_link,
        )

        # If logged-in department staff, set the department automatically
        if request.user.userprofile.is_department_staff:
            programme.department = request.user.userprofile.department

        programme.save()
        messages.success(request, "Programme added successfully.")
        return redirect("program_list")  # Redirect to the programme list view

    return render(request, "AddData/add_programme.html")
def edit_programme_view(request, programme_id):
    """
    View to edit an existing programme.
    """
    # Fetch the programme object based on ID or return a 404 if not found
    programme = get_object_or_404(Programme, id=programme_id)

    if request.method == "POST":
        # Fetch updated data from the POST request
        programme.programme_code = request.POST.get("programme_code", programme.programme_code)
        programme.programme_name = request.POST.get("programme_name", programme.programme_name)
        programme.year_of_introduction = request.POST.get("year_of_introduction", programme.year_of_introduction)
        programme.cbcs_status = request.POST.get("cbcs_status", programme.cbcs_status)
        programme.year_of_cbcs_implementation = request.POST.get("year_of_cbcs_implementation", programme.year_of_cbcs_implementation)
        programme.year_of_revision = request.POST.get("year_of_revision", programme.year_of_revision)
        programme.content_update_percentage = request.POST.get("content_update_percentage", programme.content_update_percentage)
        programme.document_link = request.POST.get("document_link", programme.document_link)

        try:
            # Save the updated programme
            programme.save()
            messages.success(request, "Programme updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating programme: {e}")

        # Redirect to the programme list view after updating
        return redirect("program_list")

    # Render the edit form with current programme data if the request is GET
    context = {
        "programme": programme,
    }
    return render(request, "Forms/programme_list.html", context)
def delete_programme_view(request, programme_id):
    """
    View to delete an existing programme.
    """
    # Fetch the programme object based on ID or return a 404 if not found
    programme = get_object_or_404(Programme, id=programme_id)

    try:
        # Delete the programme
        programme.delete()
        messages.success(request, "Programme deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting programme: {e}")

    # Redirect to the programme list view after deletion
    return redirect("program_list")
from openpyxl.styles import Font
def download_excel(request):
    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Programmes"

    # Define headers
    headers = [
        "Programme Code", "Programme Name", "Year of Introduction",
        "CBCS Status", "Year of CBCS Implementation",
        "Year of Revision", "Content Update (%)", "Document Link"
    ]
    row_num = 1

    # Add headers to the sheet
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Bold headers

    # Add data rows
    programmes = Programme.objects.all()
    for programme in programmes:
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = programme.programme_code
        worksheet.cell(row=row_num, column=2).value = programme.programme_name
        worksheet.cell(row=row_num, column=3).value = programme.year_of_introduction
        worksheet.cell(row=row_num, column=4).value = programme.cbcs_status
        worksheet.cell(row=row_num, column=5).value = programme.year_of_cbcs_implementation or "-"
        worksheet.cell(row=row_num, column=6).value = programme.year_of_revision or "-"
        worksheet.cell(row=row_num, column=7).value = programme.content_update_percentage or "-"
        worksheet.cell(row=row_num, column=8).value = programme.document_link or "-"

    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Programmes.xlsx"'
    workbook.save(response)
    return response
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from .models import Programme

def download_pdf(request):
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer, pagesize=A4, 
        rightMargin=30, leftMargin=30, 
        topMargin=30, bottomMargin=30
    )

    styles = getSampleStyleSheet()
    text_style = styles["BodyText"]
    text_style.wordWrap = "CJK"  # Proper word wrapping

    # Table headers
    data = [
        [
            "Programme Code", "Programme Name", "Year of Introduction",
            "CBCS Status", "Year of CBCS Implementation", "Year of Revision",
            "Content Update (%)", "Document Link"
        ]
    ]

    # Fetch and format programme data
    programmes = Programme.objects.all()
    for programme in programmes:
        data.append([
            Paragraph(truncate_text(str(programme.programme_code or "-"), 15), text_style),
            Paragraph(truncate_text(str(programme.programme_name or "-"), 25), text_style),
            str(programme.year_of_introduction) if programme.year_of_introduction else "-",
            Paragraph(str(programme.cbcs_status or "-"), text_style),
            str(programme.year_of_cbcs_implementation) if programme.year_of_cbcs_implementation else "-",
            str(programme.year_of_revision) if programme.year_of_revision else "-",
            str(programme.content_update_percentage) if programme.content_update_percentage else "-",
            Paragraph(f'<a href="{programme.document_link}">{truncate_text(programme.document_link or "-", 40)}</a>', styles["Normal"]),
        ])

    # Adjust table layout
    table = Table(data, colWidths=[70, 120, 60, 70, 70, 70, 80, 120])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),  # Header background
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Center alignment
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),  # Bold header font
        ("FONTSIZE", (0, 0), (-1, 0), 12),  # Header font size
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),  # Padding below header
        ("GRID", (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),  # Vertical alignment
    ]))

    # Add title and table to elements
    elements = [Paragraph("Programmes List", styles["Title"]), Spacer(1, 20), table]

    # Build PDF with a border
    pdf.build(elements, onFirstPage=add_border, onLaterPages=add_border)

    # Return PDF as response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Programmes_A4.pdf"'
    return response

def truncate_text(text, max_length):
    """ Truncates text to a maximum length and appends '...' if it exceeds. """
    return text if len(text) <= max_length else text[:max_length - 3] + "..."

def add_border(canvas, doc):
    """ Draws a border around the page content. """
    canvas.saveState()
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(1)
    canvas.rect(
        doc.leftMargin, doc.bottomMargin, 
        doc.width, doc.height
    )
    canvas.restoreState()

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4

@login_required
def download_programme_pdf(request, programme_id):
    """
    Generate and download a PDF for a specific Programme record.
    """
    # Fetch the specific Programme record
    programme = get_object_or_404(Programme, id=programme_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="programme_{programme.programme_code}.pdf"'

    # Create the PDF canvas
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # Header Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Programme Details")

    # Table Data for Programme Details
    data = [
        ["Programme Code", programme.programme_code],
        ["Programme Name", programme.programme_name],
        ["Year of Introduction", programme.year_of_introduction],
        ["CBCS Status", programme.cbcs_status],
        ["Year of CBCS Implementation", programme.year_of_cbcs_implementation or "N/A"],
        ["Year of Revision", programme.year_of_revision or "N/A"],
        ["Content Update (%)", programme.content_update_percentage or "N/A"],
        ["Document Link", programme.document_link or "N/A"],
    ]

    # Style the Table
    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # Draw Table in the Center
    table_x = (width - 480) / 2
    table_y = height - 350
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # Footer Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Page Borders**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response
def view_programme(request, id):
    """
    Displays the details of a specific programme.
    """
    # Fetch the programme object or return 404 if it doesn't exist
    programme = get_object_or_404(Programme, id=id)

    # Pass the programme to the template
    return render(request, "viewData/programme_detail.html", {"programme": programme})

@login_required
def view_courses(request):
    """
    View function for displaying courses.
    Department staff can see their department's data,
    and superusers can see all department data.
    """
    if request.user.is_superuser:
        # Superuser sees all courses
        courses = Course.objects.all()
    else:
        # Non-superuser sees courses for their department
        user_department = getattr(request.user, 'userprofile', None).department
        courses = Course.objects.filter(department=user_department)

    return render(request, "Forms/course_list.html", {"courses": courses})
@login_required
def add_course(request):
    """
    View to handle adding a new course with restrictions:
    - Department staff can only see their department in the select dropdown.
    - Superusers can see all departments.
    """
    if request.method == "POST":
        department = Department.objects.get(department_code=request.POST["department"])
        name = request.POST["name"]
        code = request.POST["code"]
        year_of_introduction = request.POST["year_of_introduction"]
        activities = request.POST["activities"]
        document = request.FILES.get("document", None)

        # Create and save the new course
        Course.objects.create(
            department=department,
            name=name,
            code=code,
            year_of_introduction=year_of_introduction,
            activities=activities,
            document=document
        )
        return redirect("view_courses")

    # Fetch departments for the dropdown
    if request.user.is_superuser:
        # Superusers can see all departments
        departments = Department.objects.all()
    else:
        # Staff can only see their associated department
        user_department = getattr(request.user, "userprofile", None).department
        departments = Department.objects.filter(department_code=user_department.department_code)

    return render(request, "AddData/add_course.html", {"departments": departments})

@login_required
def edit_course(request, course_id):
    """
    View to edit an existing course.
    """
    # Fetch the course object based on ID or return 404 if not found
    course = get_object_or_404(Course, id=course_id)

    # Check if the request method is POST
    if request.method == "POST":
        # Get updated values from the POST request
        course.name = request.POST.get("name", course.name)
        course.code = request.POST.get("code", course.code)
        course.year_of_introduction = request.POST.get("year_of_introduction", course.year_of_introduction)
        course.activities = request.POST.get("activities", course.activities)

        # Update the document if a new file is uploaded
        if "document" in request.FILES:
            course.document = request.FILES["document"]

        # Update the department only if the user is a superuser
        if request.user.is_superuser:
            department_id = request.POST.get("department", course.department.id)
            course.department = get_object_or_404(Department, id=department_id)

        try:
            # Save the updated course
            course.save()
            messages.success(request, "Course updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating course: {e}")

        # Redirect to the course list page
        return redirect("view_courses")

    # Fetch department options for the dropdown
    if request.user.is_superuser:
        departments = Department.objects.all()  # Superuser can view all departments
    else:
        departments = Department.objects.filter(id=course.department.id)  # Staff sees their own department

    # Render the form with pre-filled course data for GET requests
    context = {
        "course": course,
        "departments": departments,
    }
    return render(request, "Forms/course_list.html", context)
@login_required
def delete_course_view(request, course_id):
    """
    View to delete an existing course.
    """
    # Fetch the course object based on ID or return a 404 if not found
    course = get_object_or_404(Course, id=course_id)

    # Restrict department staff to only delete their department's courses
    if not request.user.is_superuser:
        user_department = getattr(request.user, "userprofile", None).department
        if course.department != user_department:
            messages.error(request, "You do not have permission to delete this course.")
            return redirect("view_courses")

    # Delete the course
    try:
        course.delete()
        messages.success(request, "Course deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting course: {e}")

    # Redirect to the course list after deletion
    return redirect("view_courses")
@login_required
def download_course_pdf(request, course_id):
    """
    Generate and download a PDF for a specific Course record.
    """
    # Fetch the specific Course record
    course = get_object_or_404(Course, id=course_id)

    # Set up the HTTP response for a PDF file
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="course_{course.code}.pdf"'

    # Create the PDF canvas
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # **Background Color**
    pdf.setFillColorRGB(0.95, 0.95, 0.95)
    pdf.rect(0, 0, width, height, fill=True, stroke=False)

    # Header Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, height - 100, width - 40, 80, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 60, "Course Details")

    # Table Data for Course Details
    data = [
        ["Course Name", course.name],
        ["Course Code", course.code],
        ["Year of Introduction", course.year_of_introduction],
        ["Employability/Skill Development Activities", course.activities],
        ["Department", course.department.department_name],
    ]

    # Include document link if available
    if course.document:
        data.append(["Document", course.document.url])
    else:
        data.append(["Document", "No Document Available"])

    # Style the Table
    table = Table(data, colWidths=[200, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 1, colors.blue),
        ("BOX", (0, 0), (-1, -1), 2, colors.blue),
    ]))

    # Draw Table in the Center
    table_x = (width - 480) / 2
    table_y = height - 350
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, table_x, table_y)

    # Footer Section
    pdf.setFillColor(colors.blue)
    pdf.roundRect(20, 20, width - 40, 50, 10, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, 45, "© 2025 Bharathiar University. All Rights Reserved.")

    # **Page Borders**
    pdf.setStrokeColor(colors.blue)
    pdf.setLineWidth(2)
    pdf.roundRect(10, 10, width - 20, height - 20, 15, stroke=True, fill=False)
    pdf.roundRect(15, 15, width - 30, height - 30, 12, stroke=True, fill=False)

    pdf.showPage()
    pdf.save()

    return response
@login_required
def course_detail_view(request, course_id):
    """
    View to display details of a specific course.
    """
    # Fetch the course object or raise a 404 if not found
    course = get_object_or_404(Course, id=course_id)

    # Render the course detail template
    context = {
        "course": course,
    }
    return render(request, "viewData/course_details.html", context)
def value_added_course(request):
    """
    View to display a list of Value-Added Courses.
    - Department staff can only see their department's records.
    - Superusers can see all records.
    """
    if request.user.is_superuser:
        # Superusers see all records
        courses = ValueAddedCourse.objects.all()
    else:
        # Department staff only see courses from their department
        user_department = getattr(request.user, "userprofile", None).department
        courses = ValueAddedCourse.objects.filter(department=user_department)

    context = {
        "courses": courses,
    }
    return render(request, "Forms/value_added_course.html", context)

def add_value_added_course(request):
    """
    View to add a new Value-Added Course.
    - Department staff can only add data for their own department.
    - Superusers can add data for any department.
    """
    if request.method == "POST":
        # Handle POST request
        if request.user.is_superuser:
            # Superuser can select any department
            department = get_object_or_404(Department, department_code=request.POST["department"])
        else:
            # Department staff can only select their associated department
            department = getattr(request.user, "userprofile", None).department

        name = request.POST["name"]
        code = request.POST["code"]
        year_of_offering = request.POST["year_of_offering"]
        times_offered = request.POST["times_offered"]
        duration = request.POST["duration"]
        students_enrolled = request.POST["students_enrolled"]
        students_completed = request.POST["students_completed"]
        document = request.FILES.get("document", None)

        # Create the new course
        ValueAddedCourse.objects.create(
            department=department,
            name=name,
            code=code,
            year_of_offering=year_of_offering,
            times_offered=times_offered,
            duration=duration,
            students_enrolled=students_enrolled,
            students_completed=students_completed,
            document=document
        )

        return redirect("value_added_courses")

    # Handle GET request: Display the form
    if request.user.is_superuser:
        # Superuser sees all departments
        departments = Department.objects.all()
    else:
        # Department staff only sees their own department
        departments = Department.objects.filter(department_code=request.user.userprofile.department.department_code)

    return render(request, "AddData/add_value_added_course.html", {"departments": departments})
@login_required
def edit_value_added_course(request, course_id):
    """
    View to edit an existing Value-Added Course record.
    - Department staff can only edit records for their department.
    - Superusers can edit all records.
    """
    # Fetch the course or return a 404 if not found
    course = get_object_or_404(ValueAddedCourse, id=course_id)

    # Restrict department staff to only edit their department's courses
    if not request.user.is_superuser:
        user_department = getattr(request.user, "userprofile", None).department
        if course.department != user_department:
            messages.error(request, "You do not have permission to edit this record.")
            return redirect("value_added_courses")

    if request.method == "POST":
        # Update the fields
        course.name = request.POST.get("name", course.name)
        course.code = request.POST.get("code", course.code)
        course.year_of_offering = request.POST.get("year_of_offering", course.year_of_offering)
        course.times_offered = request.POST.get("times_offered", course.times_offered)
        course.duration = request.POST.get("duration", course.duration)
        course.students_enrolled = request.POST.get("students_enrolled", course.students_enrolled)
        course.students_completed = request.POST.get("students_completed", course.students_completed)

        # Replace the document if a new file is uploaded
        if "document" in request.FILES:
            course.document = request.FILES["document"]

        try:
            course.save()
            messages.success(request, "Record updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating record: {e}")
        return redirect("value_added_courses")

    context = {
        "course": course,
    }
    return render(request, "Forms/value_added_course.html", context)
@login_required
def delete_value_added_course(request, course_id):
    """
    View to delete an existing Value-Added Course record.
    - Superusers can delete any record.
    - Department staff can only delete records from their own department.
    """
    # Fetch the course or return a 404 if not found
    course = get_object_or_404(ValueAddedCourse, id=course_id)

    # Restrict department staff to only delete their department's courses
    if not request.user.is_superuser:
        user_department = getattr(request.user, "userprofile", None).department
        if course.department != user_department:
            messages.error(request, "You do not have permission to delete this record.")
            return redirect("value_added_courses")

    try:
        course.delete()
        messages.success(request, "Record deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting record: {e}")

    return redirect("value_added_courses")


def student_project_list_view(request):
    """
    View to display a list of student projects.
    - Department staff can only see records for their department.
    - Superusers can see all records.
    """
    if request.user.is_superuser:
        # Superusers can view all records
        projects = StudentProject.objects.all()
    else:
        # Department staff can only view their department's records
        user_department = getattr(request.user, "userprofile", None).department
        projects = StudentProject.objects.filter(department=user_department)

    context = {
        "projects": projects,
    }
    return render(request, "Forms/field_project.html", context)
def add_student_project_view(request):
    """
    View to add a new student project record.
    - Department staff can only add data for their department.
    - Superusers can add data for any department.
    """
    if request.method == "POST":
        if request.user.is_superuser:
            # Superuser can select any department
            department = get_object_or_404(Department, department_code=request.POST["department"])
        else:
            # Department staff can only select their associated department
            department = getattr(request.user, "userprofile", None).department

        programme_name = request.POST["programme_name"]
        programme_code = request.POST["programme_code"]
        students = request.POST["students"]
        document = request.FILES.get("document", None)

        # Create the new record
        StudentProject.objects.create(
            department=department,
            programme_name=programme_name,
            programme_code=programme_code,
            students=students,
            document=document,
        )

        return redirect("student_project_list")

    if request.user.is_superuser:
        # Superusers see all departments
        departments = Department.objects.all()
    else:
        # Department staff only see their own department
        departments = Department.objects.filter(department_code=request.user.userprofile.department.department_code)

    return render(request, "AddData/add_field_project.html", {"departments": departments})
@login_required
def edit_student_project_view(request, project_id):
    """
    View to edit an existing student project record.
    """
    # Fetch the project or return a 404 if not found
    project = get_object_or_404(StudentProject, id=project_id)

    if request.method == "POST":
        # Update project details
        project.programme_name = request.POST.get("programme_name", project.programme_name)
        project.programme_code = request.POST.get("programme_code", project.programme_code)
        project.students = request.POST.get("students", project.students)

        # Handle document replacement only if a new document is uploaded
        if "document" in request.FILES:
            project.document = request.FILES["document"]

        # Save the updated project instance
        try:
            project.save()
            messages.success(request, "Record updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating record: {e}")

        return redirect("student_project_list")

    # For GET requests, render the edit modal with the current project data
    context = {
        "project": project,
    }
    return render(request, "Forms/field_project.html", context)
@login_required
def delete_student_project_view(request, project_id):
    """
    View to delete a student project record.
    - Superusers can delete any record.
    - Department staff can only delete records from their own department.
    """
    # Fetch the project or return a 404 if not found
    project = get_object_or_404(StudentProject, id=project_id)

    # Authorization check
    if not request.user.is_superuser:
        user_department = getattr(request.user, "userprofile", None).department
        if project.department != user_department:
            messages.error(request, "You do not have permission to delete this record.")
            return redirect("student_project_list")

    # If authorized, delete the record
    project.delete()
    messages.success(request, "Record deleted successfully.")
    return redirect("student_project_list")